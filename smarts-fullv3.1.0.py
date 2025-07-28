#By HPDS Tech Support
#===========ooOoo============
from datetime import datetime
import json
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import csv
import os
import glob
import sqlite3
import getpass
from zipfile import ZipFile
import subprocess
from openpyxl.styles import PatternFill
from pathlib import Path
import time
from collections import defaultdict
from collections import Counter
# Define required parameters
ssd_params = [
    "Reallocated_Sector_Ct",
    "Power_On_Hours",
    "Wear_Leveling_Count",
    "Used_Rsvd_Blk_Cnt_Tot",
    "Runtime_Bad_Block",
    "Reported_Uncorrect",
    "Hardware_ECC_Recovered",
    "Total_LBAs_Written",
    "Total Size Written (TB)"
]

hdd_params = [
    "Elements in grown defect list",
    "Total Uncorrected Errors",
    "Accumulated start-stop cycles",
    "Accumulated load-unload cycles"
]
micron_ssd_params = {
    "Raw_Read_Error_Rate": "1",
    "Reallocated_Sector_Ct": "5",
    "Reported_Uncorrect": "187",
    "Hardware_ECC_Recovered": "195",
    "Unused_Rsvd_Blk_Cnt_Tot": "180",
    "Total_LBAs_Written": "246"  # Unknown attribute at ID# 246
}

threshold_sata_ssd = {
    "Reallocated_Sector_Ct": "10",
    "Power_On_Hours": "5",
    "Wear_Leveling_Count": "10",
    "Used_Rsvd_Blk_Cnt_Tot": "10",
    "Runtime_Bad_Block": "10",
    "Reported_Uncorrect": "97",
    "Hardware_ECC_Recovered": "100",
    "Total Size Written (TB)": "6400"
} 

threshold_micron_ssd = {
    "Raw_Read_Error_Rate": "50",
    "Reallocated_Sector_Ct": "1",
    "Reported_Uncorrect": "0",
    "Hardware_ECC_Recovered": "0",
    "Unused_Rsvd_Blk_Cnt_Tot": "0",
    "Total Size Written (TB)": "6400"
    
}
#Part Numbers for SSD and HDD Disks (LOM prepration)
partnums = [
    {"Type": "HDD", "Interface": "SAS", "Size": "1.20 TB", "Describtion": "HPDS 1.2TB SAS 12G Enterprise 10K 2.5in HDD", "Part Number": "PD-S1025-1200"},
    {"Type": "HDD", "Interface": "SAS", "Size": "1.80 TB", "Describtion": "HPDS 1.8TB SAS 12G Enterprise 10K 2.5in HDD", "Part Number": "PD-S1025-600"},
    {"Type": "HDD", "Interface": "SAS", "Size": "2.40 TB", "Describtion": "HPDS 2.4TB SAS 12G Enterprise 10K 2.5in HDD", "Part Number": "PD-S1025-2400"},
    {"Type": "HDD", "Interface": "SAS", "Size": "2.00 TB", "Describtion": "HPDS 2TB NL-SAS 12G Enterprise 7.2K 3.5in HDD", "Part Number": "PD-NS7235-2000"},
    {"Type": "HDD", "Interface": "SAS", "Size": "4.00 TB", "Describtion": "HPDS 4TB NL-SAS 12G Enterprise 7.2K 3.5in HDD", "Part Number": "PD-NS7235-4000"},
    {"Type": "HDD", "Interface": "SAS", "Size": "6.00 TB", "Describtion": "HPDS 6TB NL-SAS 12G Enterprise 7.2K 3.5in HDD", "Part Number": "PD-NS7235-6000"},
    {"Type": "HDD", "Interface": "SAS", "Size": "8.00 TB", "Describtion": "HPDS 8TB NL-SAS 12G Enterprise 7.2K 3.5in HDD", "Part Number": "PD-NS7235-8000"},
    {"Type": "HDD", "Interface": "SAS", "Size": "12.00 TB", "Describtion": "HPDS 12TB NL-SAS 12G Enterprise 7.2K 3.5in HDD", "Part Number": "PD-NS7235-12000"},
    {"Type": "HDD", "Interface": "SATA", "Size": "1.94 TB", "Describtion": "HPDS 1.9TB SATA 6G Enterprise 2.5in SSD", "Part Number": "PD-SDDD25-1900"},
    {"Type": "HDD", "Interface": "SATA", "Size": "3.84 TB", "Describtion": "HPDS 3.8TB SATA 6G Enterprise 2.5in SSD", "Part Number": "PD-SDDD25-3800"},
    {"Type": "HDD", "Interface": "SATA", "Size": "7.68 TB", "Describtion": "HPDS 7.6TB SATA 6G Enterprise 2.5in SSD", "Part Number": "PD-SDDD25-7600"},
    {"Type": "HDD", "Interface": "SAS", "Size": "16.0 TB", "Describtion": "HPDS 16TB SAS 7.2K Enterprise HDD SX for SAB-HB", "Part Number": "DHBD-HS07SX-16000"},
    {"Type": "SSD", "Interface": "SATA", "Size": "960 GB", "Describtion": "HPDS 960GB SATA 6G Enterprise 2.5in SSD", "Part Number": "PD-SDDD25-960"},
    {"Type": "SSD", "Interface": "SATA", "Size": "1.92 TB", "Describtion": "HPDS 1.9TB SATA 6G Enterprise 2.5in SSD", "Part Number": "PD-SDDD25-1900"},
    {"Type": "SSD", "Interface": "SATA", "Size": "3.84 TB", "Describtion": "HPDS 3.8TB SATA 6G Enterprise 2.5in SSD", "Part Number": "PD-SDDD25-3800"},
    {"Type": "SSD", "Interface": "SATA", "Size": "7.68 TB", "Describtion": "HPDS 7.6TB SATA 6G Enterprise 2.5in SSD", "Part Number": "PD-SDDD25-7600"},
    {"Type": "SSD", "Interface": "SAS", "Size": "960 GB", "Describtion": "HPDS 960GB SAS 6G Enterprise 2.5in SSD", "Part Number": "DHBD-SSXXGG-960"},
    {"Type": "SSD", "Interface": "SAS", "Size": "1.92 TB", "Describtion": "HPDS 1.9TB SAS 6G Enterprise 2.5in SSD", "Part Number": "DHBD-SSXXGG-1920"},
    {"Type": "SSD", "Interface": "SAS", "Size": "3.84 TB", "Describtion": "HPDS 3.8TB SAS 6G Enterprise 2.5in SSD", "Part Number": "DHBD-SSXXGG-3840"},
    {"Type": "SSD", "Interface": "SAS", "Size": "7.68 TB", "Describtion": "HPDS 7.6TB SAS 6G Enterprise 2.5in SSD", "Part Number": "DHBD-SSXXGG-7680"}
]
partnums_card = [
    {"Type": "FC", "Ports": 2, "Speed": "8Gb", "Describtion": "HPDS 8Gb 2-port PCIe Fiber Channel HBA", "Part Number": "PI-FM08G2P"},
    {"Type": "FC", "Ports": 4, "Speed": "8Gb", "Describtion": "HPDS 8Gb 4-port PCIe Fiber Channel HBA", "Part Number": "PI-FM08G4P"},
    {"Type": "FC", "Ports": 2, "Speed": "16Gb", "Describtion": "HPDS 16Gb 2-port PCIe Fiber Channel HBA", "Part Number": "PI-FM16G2P"},
    {"Type": "FC", "Ports": 4, "Speed": "16Gb", "Describtion": "HPDS 16Gb 4-port PCIe Fiber Channel HBA", "Part Number": "PI-FM16G4P"},
    {"Type": "NIC", "Ports": 4, "Speed": "1Gb", "Describtion": "HPDS 1Gb 4-port PCIe Ethernet Adapter", "Part Number": "PI-N1GT4P"},
    {"Type": "NIC", "Ports": 2, "Speed": "10Gb", "Describtion": "HPDS 10Gb 2-port PCIe Ethernet Adapter", "Part Number": "PI-NT10G2P"}
]
smart_pattern = re.compile(
    r"(\d+)\s+([\w_]+)\s+0x[0-9a-fA-F]+\s+(\d+)\s+\d+\s+(\d+)\s+\w+-?\w*\s+\w+\s+-\s+(\d+)"
)
# Function to extract SSD parameters
def extract_ssd_parameters(log_content):
    data = []
    # Split the log into blocks for each disk.
    disk_blocks = re.findall(
        r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)",
        log_content, re.DOTALL
    )
    
    for block in disk_blocks:
        # Determine disk type based on certain strings
        is_ssd = not re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        is_sas_ssd = re.search(r"Transport protocol:\s+SAS", block, re.IGNORECASE)
        is_micron_ssd = re.search(r"Device Model:\s+Micron", block, re.IGNORECASE)
        disk_type = ""
        capacity_match = re.search(r"\[([^\]]+)\]", block)
        if capacity_match:
            tb_value = capacity_match.group(1)
            user_capacity = float(re.search(r'\d+(?:\.\d+)?', tb_value).group()) if re.search(r'\d+(?:\.\d+)?', tb_value) else 0            
        if is_ssd:
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            model_match =  re.search(r"Device Model:\s+(.+)", block, re.IGNORECASE)
            device_model = model_match.group(1) if model_match else "Unknown" 
            if is_sas_ssd:
                # --- SAS SSD extraction (existing logic) ---
                elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
                start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
                load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)
                read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
                write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
                endurance_indicator = re.search(r"SS Media used endurance indicator:\s+(\d+)", block)
                model_match_sas = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)
                device_model = model_match_sas.group(1) if model_match_sas else "Unknown" 
                brand = "Samsung"
                disk_type = "SSD SAS"
                read_error_value = int(read_error_match.group(1)) if read_error_match else 0
                write_error_value = int(write_error_match.group(1)) if write_error_match else 0
                total_uncorrected_errors = read_error_value + write_error_value
                total_uncorrrected_threshold = int(user_capacity)  * 5
                Element_in_defect_threshold = int(user_capacity)  * 100
        #We convert to str again for consistency
                threshold_sas_ssd = {
                    "Elements in grown defect list": str(Element_in_defect_threshold),
                    "Total Uncorrected Errors": str(total_uncorrrected_threshold),
                    "Accumulated start-stop cycles": "10000",
                    "Accumulated load-unload cycles": "300000",
                    "SS Media used endurance indicator %": "90"
                    }
                hdd_values = [
                    ("Elements in grown defect list", elements_grown_defect),
                    ("Total Uncorrected Errors", total_uncorrected_errors),
                    ("Accumulated start-stop cycles", start_stop_cycles),
                    ("SS Media used endurance indicator %", endurance_indicator)
                ]

                for param, match in hdd_values:
                    if isinstance(match, int):  # For Total Uncorrected Errors
                        raw_value = str(match)
                    elif match:
                        raw_value = match.group(1)
                    else:
                        continue
                    threshold = threshold_sas_ssd.get(param, "-")
                    data.append({
                        "Brand" : brand,
                        "Device Model": device_model, 
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": param,
                        "Threshold" : threshold,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
                    
            else:
                # --- Non-SAS SSD extraction ---
                # smart_pattern should be defined elsewhere to extract tuples:
                # (attr_id, attr_name, value, raw_value)
                smart_matches = smart_pattern.findall(block)
                total_lba_written = None  # To store Total_LBAs_Written when found
                disk_type = "SSD SATA"
                if is_micron_ssd:
                    brand = "Micron"
                    # Define the list of expected Micron SMART parameters.
                    micron_params = [
                        "Raw_Read_Error_Rate", 
                        "Reallocated_Sector_Ct", 
                        "Reported_Uncorrect", 
                        "Hardware_ECC_Recovered", 
                        "Unused_Rsvd_Blk_Cnt_Tot",
                        "Used_Reserve_Block_Count"
                    ]

                    for match in smart_matches:
                        attr_id, attr_name, value, threshold, raw_value = match
                        # Check for the known Micron parameters.
                        if attr_name in micron_params:
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": attr_name,
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        # Detect Total_LBAs_Written by its attribute ID "246"
                        elif attr_id.strip() == "246":
                            try:
                                total_lba_written = int(raw_value)
                            except ValueError:
                                total_lba_written = None
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": "Total_LBAs_Written",
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        #Detect used block count
                        elif attr_id.strip() == "170":
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": "Used_Reserve_Block_Count",
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        elif attr_id.strip() == "202":
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": "Percent Life Time Remaining",
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                    if total_lba_written is not None:

                        total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                        threshold_write_micron = 365 * 0.8 * 5 * user_capacity
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)",
                            "Threshold": str(round(threshold_write_micron, -3)),
                            "Value": "-",
                            "Raw Value": f"{total_size_written_tb:.2f}"
                        })

                        total_size_micron_waf2 = total_size_written_tb * 2
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-WAF(2)",
                            "Threshold": str(round(threshold_write_micron, -1)),
                            "Value": "-",
                            "Raw Value": f"{total_size_micron_waf2:.2f}"
                        })
                        total_size_micron_waf4 = total_size_written_tb * 4
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-WAF(4)",
                            "Threshold": str(round(threshold_write_micron, -1)),
                            "Value": "-",
                            "Raw Value": f"{total_size_micron_waf4:.2f}"
                        })

                else:
                    brand = "SAMSUNG"
                    disk_type = "SSD SATA"
                    # Existing logic for non-Micron SATA SSDs
                    for match in smart_matches:
                        attr_id, attr_name, value, threshold, raw_value = match
                        if attr_name in ssd_params: # ssd_params defined elsewhere
                            if attr_name == "Total_LBAs_Written":
                                try:
                                    total_lba_written = int(raw_value)
                                except ValueError:
                                    total_lba_written = None
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": attr_name,
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        if attr_name == "Wear_Leveling_Count":
                            wearlevel = raw_value
                    if total_lba_written is not None:
                        total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                        if "MZ" in device_model:
                            threshold_samsung_write = round(365 * 5 * user_capacity * 4, -3)
                        elif "850" in device_model:
                            threshold_samsung_write = 150
                            user_capacity = 0.25
                        elif "860" in device_model:
                            threshold_samsung_write = 300
                            user_capacity = 0.25
                        elif "870" in device_model:
                            threshold_samsung_write = 150
                            user_capacity = 0.25
                        else:
                            threshold_samsung_write = 300 #undefined
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)",
                            "Threshold": str(threshold_samsung_write),
                            "Value": "-",
                            "Raw Value": f"{total_size_written_tb:.2f}"
                        })
                        total_size_written_ssd = float(wearlevel) * user_capacity
                        total_size_samsung_waf2 = total_size_written_tb * 2
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-WAF(2)",
                            "Threshold": str(round(threshold_samsung_write, -2)),
                            "Value": "-",
                            "Raw Value": f"{total_size_samsung_waf2:.2f}"
                        })
                        total_size_samsung_waf4 = total_size_written_tb * 4
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-WAF(4)",
                            "Threshold": str(round(threshold_samsung_write, -2)),
                            "Value": "-",
                            "Raw Value": f"{total_size_samsung_waf4:.2f}"
                        })
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-SSD_Ctl",
                            "Threshold": str(round(threshold_samsung_write, -2)),
                            "Value": "-",
                            "Raw Value": f"{total_size_written_ssd:.2f}"
                        })
            # Add an empty row after each disk's data for readability.
            data.append({
                "Brand": "",
                "Device Model": "",
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
            for i in range(len(data) - 2):
                if (
                    data[i]["Parameter"] == "Total_LBAs_Written" and
                    data[i + 1]["Parameter"] == "Unused_Rsvd_Blk_Cnt_Tot"
                ):
                    # Swap their positions
                    data[i], data[i + 1] = data[i + 1], data[i]
            
    return data

# Function to extract HDD parameters
def extract_hdd_parameters(log_content):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    for block in disk_blocks:
        capacity_match = re.search(r"\[([^\]]+)\]", block)
        if capacity_match:
            tb_value = capacity_match.group(1)
        user_capacity = float(re.search(r'\d+(?:\.\d+)?', tb_value).group()) if re.search(r'\d+(?:\.\d+)?', tb_value) else 0
        total_uncorrrected_threshold = int(user_capacity)  * 5
        Element_in_defect_threshold = int(user_capacity)  * 100
        #We convert to str again for consistency
        threshold_hdd_sata = {
                "Elements in grown defect list": str(Element_in_defect_threshold),
                "Total Uncorrected Errors": str(total_uncorrrected_threshold),
                "Accumulated start-stop cycles": "10000",
                "Accumulated load-unload cycles": "300000"
                }
        is_sata = re.search(r'\bSATA\b', block, re.IGNORECASE)         
        is_hdd = re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        if is_hdd and not is_sata:

            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            disk_type = "HDD SAS"
            model_match =  re.search(r"Device Model:\s+(.*?)\s*$", block, re.IGNORECASE)
            model_match_hp = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)
            device_model = (
                    model_match.group(1) if model_match 
                    else model_match_hp.group(1) if model_match_hp
                    else "Unknown"
                    )
            #seagate_match =  re.search(r"Vendor:\s+(\S+)", block, re.IGNORECASE)
            if device_model.startswith("ST"):
                brand = "SEAGATE"
            elif device_model.startswith("TOSHIBA"):
                brand = "Toshiba"
            else:
                brand = "HP"

  
            elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
            start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
            load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)

            # Extract Total Uncorrected Errors
            read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
            write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
            
            read_error_value = int(read_error_match.group(1)) if read_error_match else 0
            write_error_value = int(write_error_match.group(1)) if write_error_match else 0
            total_uncorrected_errors = read_error_value + write_error_value

            # Sort the parameters in the desired order
            hdd_values = [
                ("Elements in grown defect list", elements_grown_defect),
                ("Total Uncorrected Errors", total_uncorrected_errors),
                ("Accumulated start-stop cycles", start_stop_cycles),
                ("Accumulated load-unload cycles", load_unload_cycles)
            ]
            
            for param, match in hdd_values:
                if isinstance(match, int):  # For Total Uncorrected Errors
                    raw_value = str(match)
                    threshold = threshold_hdd_sata.get(param, "-")
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": param,
                        "Threshold": threshold,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
                elif match:
                    raw_value = match.group(1) if hasattr(match, "group") else str(match)
                    threshold = threshold_hdd_sata.get(param, "-")
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": param,
                        "Threshold": threshold,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
            # Add an empty row after each disk's data
            data.append({
                "Brand": "",
                "Device Model": "",
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
        elif is_hdd and is_sata:
            # --- SATA HDD extraction ---
            # smart_pattern should be defined elsewhere to extract tuples:
            # (attr_id, attr_name, value, raw_value)
            smart_matches = smart_pattern.findall(block) 
            total_lba_written = None  # To store Total_LBAs_Written when found    
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            disk_type = "HDD SATA"
            model_match = re.search(r"Device Model:\s+(.*?)\s*$", block, re.IGNORECASE | re.MULTILINE)
            model_match_hp = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)

            #seagate_match =  re.search(r"Vendor:\s+(\S+)", block, re.IGNORECASE)
            #brand = seagate_match.group(1) if seagate_match else "HP"
            device_model = (
                    model_match.group(1) if model_match 
                    else model_match_hp.group(1) if model_match_hp
                    else "Unknown"
                    )
            if device_model.startswith("ST"):
                brand = "SEAGATE"
            elif device_model.startswith("TOSHIBA"):
                brand = "Toshiba"
            # Define the list of expected Micron SMART parameters.
            hdd_sata_params = [
                "Raw_Read_Error_Rate", 
                "Spin_Retry_Count", 
                "Reported_Uncorrect",
                "Command_Timeout", 
                "Current_Pending_Sector",
                "Offline_Uncorrectable",
                "Multi_Zone_Error_Rate"
            ]
            for match in smart_matches:
                attr_id, attr_name, value, threshold, raw_value = match
                # Check for the known SATA parameters.
                if attr_name in hdd_sata_params:
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": attr_name,
                        "Threshold": threshold,
                        "Value": value,
                        "Raw Value": raw_value
                    })
                    # Detect Total_LBAs_Written by its attribute ID "241"
                elif attr_id.strip() == "241":
                    try:
                        total_lba_written = int(raw_value)
                    except ValueError:
                        total_lba_written = None
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": "Total_LBAs_Written",
                        "Threshold": threshold,
                        "Value": value,
                        "Raw Value": raw_value
                    })
                    
            if total_lba_written is not None:
                total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                threshold = threshold_micron_ssd.get("Total Size Written (TB)", "-")
                data.append({
                    "Brand": brand,
                    "Device Model": device_model,
                    "Serial Number": serial_number,
                    "Interface": disk_type,
                    "Size": tb_value,
                    "Parameter": "Total Size Written (TB)",
                    "Threshold": threshold,
                    "Value": "-",
                    "Raw Value": f"{total_size_written_tb:.2f}"
                    })

            data.append({
                "Brand": "",
                "Device Model": "",
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
    return data

# Function to extract HDD device info (only if both values are found)
def extract_device_info(log_content, enclosure):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    tmp_data = []
    for block in disk_blocks:
         serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
         serial_number = serial_match.group(1) if serial_match else None

         temp_match = re.search(r"Current Drive Temperature:\s+(\d+)", block)
         hours_match = re.search(r"number of hours powered up\s+=\s+([\d.]+)", block)
         #For Samsung SSDs
         temp_match_sam = re.search(r"194\s+[\w_]+\s+[\w\d]+\s+\d+\s+\d+\s+\d+\s+\w+\s+\w+\s+-\s+(\d+)", block)
         if not temp_match_sam: temp_match_sam = re.search(r"190\s+[\w_]+\s+[\w\d]+\s+\d+\s+\d+\s+\d+\s+\w+\s+\w+\s+-\s+(\d+)", block) 
         hours_match_sam = re.search(r"9\s+[\w_]+\s+[\w\d]+\s+\d+\s+\d+\s+\d+\s+\w+\s+\w+\s+-\s+(\d+)", block)
         tsw = "-"
         tsw_ctl = "-"
         hours = "-"
         tsw_waf2 = "-"
         tsw_waf4 = "-"
         percent_life = "-"
         if serial_number:
             if temp_match:
                 temperature = f"{temp_match.group(1)}"
             else:
                 temperature = "-"

             for enc in enclosure:
                 if serial_number == enc["Serial Number"]:
                     slot = enc["En/Slot"]
                     model = enc["Device Model"]
                     if enc["Parameter"] == "Total Size Written (TB)":
                         tsw = enc["Raw Value"]
                         tsw_waf2 = float(tsw) * 2
                         tsw_waf4 = float(tsw) * 4
                     elif enc["Parameter"] == "Total Size Written (TB)-SSD_Ctl":
                         tsw_ctl = enc["Raw Value"]
                     if enc["Parameter"] == "Power_On_Hours":
                         hours = enc["Raw Value"]
                     if enc["Parameter"] == "Percent Life Time Remaining":
                         percent_life = enc["Raw Value"]
             if hours_match:
                 hours = hours_match.group(1)
             tmp_data.append({
                "Enc/Slot": slot,
                "Model": model,
                "Serial Number": serial_number,
                "Temperature": temperature,
                "Powered Up Hours": hours,
                "TBW": tsw,
                "TBW-WAF2": tsw_waf2,
                "TBW-WAF4": tsw_waf4,
                "TBW (Ctl)": tsw_ctl,
                "Percent Life": percent_life
                })

         elif serial_number:
             if temp_match_sam:
                temperature = f"{temp_match_sam.group(1)}"
             else:
                temperature = "-"
             if hours_match_sam:
                 hours = hours_match_sam.group(1)
             data.append({
                 "Enc/Slot": slot,
                 "Model": model,
                 "Serial Number": serial_number,
                 "Temperature": temperature,
                 "Powered Up Hours": hours,
                 "TBW": tsw,
                 "TBW-WAF2": tsw_waf2,
                 "TBW-WAF4": tsw_waf4,
                 "TBW (Ctl)": tsw_ctl,
                 "Percent Life": percent_life
                 })

    data.extend(tmp_data)
    return data

# Function to extract enclosure/slot information
def extract_enclosure_slot_info(log_content, serial_numbers):
    enclosure_slot_data = {}

    # Split the log content into lines
    lines = log_content.splitlines()

    current_shield_counter = None
    current_media_error_count = None
    current_other_error_count = None
    predictive_failure_count = None
    disk_status = None
    for i, line in enumerate(lines):
        line = line.strip()

        # Capture Shield Counter
        if line.startswith("Shield Counter"):
            current_shield_counter = line.split("=")[1].strip()
            if current_shield_counter == "0":
                current_shield_counter = "-"

        # Capture Media Error Count
        elif line.startswith("Media Error Count"):
            current_media_error_count = line.split("=")[1].strip()
            if current_media_error_count == "0":
                current_media_error_count = "-"


        # Capture Other Error Count
        elif line.startswith("Other Error Count"):
            current_other_error_count = line.split("=")[1].strip()
            if current_other_error_count == "0":
                current_other_error_count = "-"

        #Capture Predictive Failure Count
        elif line.startswith("Predictive Failure Count"):
            predictive_failure_count = line.split("=")[1].strip()
            if predictive_failure_count == "0":
               predictive_failure_count = "-"

        #Capture disk state 
        
        if " UGood " in line:
            disk_status = "No Config"
        elif " Onln " in line:
            disk_status = "Operational"
        elif " DHS " in line:
            disk_status = "Hotspare"
        elif " UBad " in line:
            disk_status = "BAD UNCONFIGURED"


        # Detect Serial Number
        elif line.startswith("SN ="):
            serial = line.split("=")[1].strip()
            if serial in serial_numbers:
                # Search backward for Drive line
                for j in range(i, max(i - 20, -1), -1):
                    drive_line = lines[j].strip()
                    if drive_line.startswith("Drive /c"):
                        enclosure = drive_line.split("/e")[1].split("/")[0]
                        slot = drive_line.split("/s")[1].split()[0]

                        # Now assign all gathered info
                        enclosure_slot_data[serial] = {
                            "enclosure_slot": f"{enclosure}/{slot}",
                            "shield_counter": current_shield_counter,
                            "media_error_count": current_media_error_count,
                            "other_error_count": current_other_error_count,
                            "predictive_failure_count": predictive_failure_count,
                            "Disk State": disk_status
                        }
                        break

    return enclosure_slot_data
def extract_sysinfo():
    sysinfo_file = os.path.join(script_dir, "SystemOverallInfo", "SystemInfo.mylinux")
    version_file = os.path.join(script_dir, "version")
    pmc_file = glob.glob(os.path.join(script_dir, "output.txt"))
    sys_info = {}
    voltage_index = 1 #For separting the two power modules
    current_index = 1 #For separating the two power modules
    try:
        if pmc_file:
            # Extract versions from pmc output
            with open(pmc_file[0], "r") as file:
                content = file.read()
                basic_info = {
                    "SAB ID": re.search(r'hostname\s*:\s*(.*)$', content, re.IGNORECASE | re.MULTILINE), 
                    "SAB Version": re.search(r'#SAB version\s+([^\s]+)', content),
                    "Replication Version": re.search(r'REPLICATION VERSION:\s*VERSION=([^\s]+)', content, re.IGNORECASE),
                    "Rapidtier Version": re.search(r'Rapidtier Version:\s*([^\s]+)', content),
                    "UI Version": re.search(r'version\s*=\s*"([^"]+)"', content),
                    "CLI Version": re.search(r'CLI Version:\s*([^\s]+)', content),
                    "ROC Temp": re.search(r'ROC temperature.*\(Degree Celsius\)\s*=\s*([^\s]+)', content),
                    "CV Temp": re.search(r'Temperature\s*([^\s]+)', content),
                    "BBU Status": re.search(r'BBU Status =\s*([^\s]+)', content)
                } 
        #If pmc file does not include desired values
            if basic_info["UI Version"] == None:
                with open(version_file, "r") as file:
                    content = file.read()
                    basic_info = {
                        "UI Version": re.search(r'UI Version:\s*([\d.]+)', content)
                    }
        else:
            # Extract versions from version file
            with open(version_file, "r") as file:
                content = file.read()
                basic_info = {
                    "UI Version": re.search(r'UI Version:\s*([\d.]+)', content),
                    "CLI Version": re.search(r'CLI Version:\s*([\d.]+)', content),
                    "SAB Version": re.search(r'SAB Version:\s*([\d.]+)', content)
                }

            
        for name, match in basic_info.items():
            if match:
                sys_info[name] = match.group(1)
            else:
                sys_info[name] = "Not Found"
        if "BBU Status" in sys_info:
           bbu_value = "Unknown" if sys_info["BBU Status"] is None else sys_info["BBU Status"]
           sys_info["BBU Status"] = "OK" if bbu_value == "0" else "Not OK"
        # Extract uptime and serial number from SystemInfo.mylinux
        with open(sysinfo_file, "r") as file:
            for line in file:
                uptime_match = re.search(r"up\s+(\d+)\s+days?", line)
                serial_match = re.search(r"Serial Number:\s*ZM(\S+)", line)
                voltage_match = re.search(r"Input Voltage\s*\|\s*([\d.]+)\s*V", line)
                current_match = re.search(r"Input Current \s*\|\s*([\d.]+)\s*A", line) 
                if uptime_match:
                    sys_info["Uptime (days)"] = int(uptime_match.group(1))
                    break
                else:
                    sys_info["Uptime (days)"] = 0
                if serial_match:
                     sys_info["Serial Number"] = serial_match.group(1)
                if voltage_match:
                    sys_info[f"Voltage{voltage_index}"] = float(voltage_match.group(1))
                    voltage_index += 1
                if current_match:
                    sys_info[f"Current{current_index}"] = float(current_match.group(1))
                    current_index += 1
        # Convert to list of dict format that pandas expects
        return [sys_info]
    except Exception as e:
        print(f"Error extracting system info: {e}")
        return []

# Function to extract host information from SCST configuration
def extract_host_info():
    scst_dir = "./SCST"
    db_dir = "./Database"
    # Checks for new scst file inside script directory
    new_scst_matches = glob.glob(os.path.join(script_dir, "scst.*"))
    if new_scst_matches:
        scst_files = new_scst_matches[0]
        input_file = scst_files
    else:
        scst_files = sorted(glob.glob(os.path.join(scst_dir, "scst_20*.conf")), reverse=True)
        input_file = scst_files[0] if scst_files else None
    
    if not input_file:
        print("Error: No 'scst_*.conf' files found in /SCST directory.")
        return []

    # Checks for output.txt file inside script directory 
    is_pmc = os.path.isfile(os.path.join(os.path.dirname(__file__), 'output.txt'))
    #print(f"[DEBUG] SCST file used: {input_file}")
    if is_pmc:
        #print("PMC output found")
        pmc_output = "output.txt"
        target_port_type = {}
        current_wwn = None
        with open(pmc_output, "r") as file:
            lines = file.readlines()
        # We map port connection to wwn addresses 
        for line in lines:
            line = line.strip()
            if line.lower().startswith("wwn = 0x"):
                hex_str = line.split('=')[1].strip().lower().replace('0x', '')
                current_wwn = ':'.join(hex_str[i:i+2] for i in range(0, len(hex_str), 2))
            elif ('Point' in line or 'NPort' in line) and current_wwn:
                port_type = "Point to Point" if 'Point' in line else "SAN Switch"
                target_port_type[current_wwn] = port_type
                current_wwn = None
    else:
        target_port_type = {}

    try:
        with open(input_file, "r") as file:
            lines = file.readlines()

        target_blocks = []
        current_target = None
        brace_count = 0

        # Manually parse TARGET blocks
        for line in lines:
            target_match = re.match(r"\s*TARGET\s+([0-9a-fA-F:]+)\s*\{", line)
            if target_match:
                if current_target:
                    print("[WARNING] Nested TARGET found, skipping previous unfinished block.")
                current_target = {
                    "address": target_match.group(1),
                    "content": [line]
                }
                brace_count = 1
                continue

            if current_target:
                current_target["content"].append(line)
                brace_count += line.count("{")
                brace_count -= line.count("}")
                if brace_count == 0:
                    # Complete block
                    target_blocks.append((current_target["address"], "".join(current_target["content"])))
                    current_target = None

        host_data = []

        for target_address, target_body in target_blocks:
            groups = re.findall(r"GROUP\s+([\w-]+)\s*\{([\s\S]*?)\}", target_body)

            for group_name, group_content in groups:
                # Handle empty Access Control
                access_control = group_name if group_name else "-"
                
                luns = re.findall(r"LUN\s+(\d+)\s+([\w_]+)", group_content)
                initiators = re.findall(r"INITIATOR\s+([0-9a-fA-F:]+)", group_content)

                unique_luns = sorted(set(luns), key=lambda x: int(x[0]))
                unique_initiators = sorted(set(initiators))

                sab_db_file = os.path.join(db_dir, "sab.db")
                host_map = {}

                if os.path.exists(sab_db_file):
                    with sqlite3.connect(sab_db_file) as conn:
                        cursor = conn.cursor()
                        for initiator in unique_initiators:
                            cursor.execute("""
                                SELECT initiator_name, to_host_id 
                                FROM hostinitiators
                                WHERE initiator_name = ?
                            """, (initiator,))
                            to_host_row = cursor.fetchone()

                            if to_host_row:
                                to_host_id = to_host_row[1]
                                cursor.execute("""
                                    SELECT NAME 
                                    FROM host
                                    WHERE ID = ?
                                """, (to_host_id,))
                                host_row = cursor.fetchone()
                                host_map[initiator] = host_row[0] if host_row else "Not Found"
                            else:
                                host_map[initiator] = "Not Found"

                # Group initiators and LUNs by host
                host_luns_initiators = {}
                for initiator, host_name in host_map.items():
                    if host_name not in host_luns_initiators:
                        host_luns_initiators[host_name] = {
                            "luns": set(),
                            "initiators": [],
                            "target_map": {}
                        }
                    host_luns_initiators[host_name]["luns"].update(
                        [lun[1] for lun in unique_luns if lun[1] != "device_null"]
                    )
                    host_luns_initiators[host_name]["initiators"].append(initiator)
                    host_luns_initiators[host_name]["target_map"][initiator] = target_address

                for host_name, data in host_luns_initiators.items():
                    host_name = host_name if host_name else "-"
                    for initiator in sorted(data["initiators"]):
                        initiator = initiator if initiator else "-"
                        target = data["target_map"].get(initiator, "-")
                        port_type = target_port_type.get(target, "-") if target_port_type else "-"
                        
                        if not data["luns"]:
                            host_data.append({
                                "Access Control": access_control,
                                "Host": host_name,
                                "LUNs": "-",
                                "Initiator Addresses": initiator,
                                "Target Address": target,
                                "Connection Type": port_type
                            })
                        else:
                            for lun in sorted(data["luns"]):
                                host_data.append({
                                    "Access Control": access_control,
                                    "Host": host_name,
                                    "LUNs": lun,
                                    "Initiator Addresses": initiator,
                                    "Target Address": target,
                                    "Connection Type": port_type
                                })

        if not host_data:
            if target_port_type:
                for target, port_type in target_port_type.items():
                    host_data.append({
                        "Access Control": "-",
                        "Host": "-",
                        "LUNs": "-",
                        "Initiator Addresses": "-",
                        "Target Address": target,
                        "Connection Type": port_type
                    })
            else:
                host_data.append({
                    "Access Control": "-",
                    "Host": "-",
                    "LUNs": "-",
                    "Initiator Addresses": "-",
                    "Target Address": "-",
                    "Connection Type": "-"
                })

        # Remove duplicates and group by all keys except "Target Address"
        def get_group_key(d):
            return tuple((k, v) for k, v in d.items() if k != "Target Address")
        grouped = defaultdict(list)
        for item in host_data:
            key = get_group_key(item)
            grouped[key].append(item["Target Address"])

        # Merging with controlled key order
        merged = []
        for key, addresses in grouped.items():
            # Rebuild original dict (without Target Address)
            new_dict = dict(key)

            # Get all keys except "Target Address"
            original_keys = list(new_dict.keys())

            # Insert "Target Address" before the last key
            if len(original_keys) >= 1:
                last_key = original_keys[-1]
                init_key = original_keys[3]
                # Remove last key temporarily
                last_value = new_dict.pop(last_key)
                init_value = new_dict.pop(init_key)
                # Add Target Address
                new_dict["Target Address"] = "__ ".join(addresses)
                # Restore last key
                new_dict[init_key] = init_value
                new_dict[last_key] = last_value
            else:
                # Edge case: only "Target Address" existed
                new_dict["Target Address"] = "__ ".join(addresses)

            merged.append(new_dict)
        return merged


    except Exception as e:
        print(f"[ERROR] extract_host_info failed: {e}")
        return []

    except ValueError as e:
        print(f"Error: {e}")
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        if 'conn' in locals():
            conn.close()
#convert wwn
def convert_wwn_hex_to_colon_format(wwn_hex: str) -> str:
    """Convert WWN from hex string like '0x51402ec001c676bc' to colon-separated format."""
    hex_clean = wwn_hex.lower().replace("0x", "")
    if len(hex_clean) != 16:
        return wwn_hex  # Return as-is if not a valid WWN length
    return ":".join(hex_clean[i:i+2] for i in range(0, 16, 2))
#Extract slot number
def extract_slot_port_info():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pmc_files = glob.glob(os.path.join(script_dir, "output.txt"))

    if pmc_files:

        with open(pmc_files[0], "r") as f:
            log = f.read()

        slot_data = defaultdict(lambda: {
            'enabled_ports': [],
            'total_ports': 0,
            'port_type': '',
            'ports': {}
        })
    
        current_slot = None
        current_type = None
        current_port = None

        for line in log.splitlines():
            line = line.strip()

            # Detect slot type
            if "--------ISCSI CARDS SLOTS" in line:
                current_type = 'iscsi'
            elif "--------FC HBA CARDS SLOTS" in line:
                current_type = 'fc'

            # Match slot/port headers
            match = re.search(r'(NIC|FC) CARD in SLOT\s+(\d+)\s+PORT\s+(\d+)(?:\s+\([^)]+\))?', line)
            if match:
                current_slot = match.group(2)
                current_port = int(match.group(3))
                slot_info = slot_data[current_slot]
                slot_info['port_type'] = current_type
                slot_info['total_ports'] += 1
                slot_info['ports'][current_port] = {
                    'wwn': '',
                    'connection_type': '',
                    'speed': '-'
                }
                continue

            # FC ports
            if current_type == 'fc':
                if line.startswith("port_speed"):
                    speed = line.split("=", 1)[1].strip()
                    slot_data[current_slot]['ports'][current_port]['speed'] = speed if "Unknown" not in speed else "-"
                elif line.startswith("wwn ="):
                    wwn = line.split("=", 1)[1].strip().lower().replace("0x", "")
                    wwn_formatted = ":".join(wwn[i:i+2] for i in range(0, len(wwn), 2))
                    slot_data[current_slot]['ports'][current_port]['wwn'] = wwn_formatted
                elif line.startswith("port_type"):
                    port_type = line.split("=", 1)[1].strip()
                    if "NPort" in port_type:
                        slot_data[current_slot]['ports'][current_port]['connection_type'] = "SAN-Switch"
                    else:
                        slot_data[current_slot]['ports'][current_port]['connection_type'] = "Point-to-Point"

            # iSCSI ports
            elif current_type == 'iscsi':
                if line.startswith("speed_interface"):
                    speed = line.split("=", 1)[1].strip()
                    slot_data[current_slot]['ports'][current_port]['speed'] = speed if "Unknown" not in speed else "-"
                elif line.startswith("mac_address"):
                    mac = line.split("=", 1)[1].strip()
                    slot_data[current_slot]['ports'][current_port]['mac_address'] = mac
                elif line.startswith("iqn."):
                    iqn = line.strip()
                    slot_data[current_slot]['ports'][current_port]['wwn'] = iqn

        result = []
        for slot, info in slot_data.items():
            result.append({
                'slot': f"Slot{slot}",
                'port_type': info['port_type'],
                'total_ports': info['total_ports'],
                'ports': info['ports']
            })

        return result
#For the purpose of counting disk type and size for LOM
def consolidate_by_serial(smart_list):
    consolidated = {}

    for entry in smart_list:
        serial = entry.get("Serial Number", "").strip()
        if not serial:
            continue  # Skip empty or malformed entries

        if serial not in consolidated:
            # Initialize with static fields
            consolidated[serial] = {
                "Serial Number": serial,
                "Brand": entry.get("Brand", ""),
                "Device Model": entry.get("Device Model", ""),
                "Interface": entry.get("Interface", ""),
                "Size": entry.get("Size", "")
            }

        param = entry.get("Parameter", "").strip()
        raw_val = entry.get("Raw Value", "").strip()

        if param:
            consolidated[serial][param] = raw_val

    return list(consolidated.values())
def lom_disk(disk_dicts):
    type_size_list = []
    disk_data = consolidate_by_serial(disk_dicts)

    for d in disk_data:
        interface = d.get("Interface", "").strip()
        size = d.get("Size", "").strip()
        model = d.get("Device Model", "").strip()
        if interface and size and type:
            disk_type = interface.split()[0]  # Get "HDD" or "SSD"
            disk_interface = interface.split()[-1]
            type_size_list.append((disk_type, size, disk_interface))

    counts = Counter(type_size_list)
    lom_disk_count = []
    for (Type, size, interface), count in counts.items():
        for part in partnums:
            if  part["Type"] == Type and part["Size"] == size and part["Interface"] == interface:
                lom_disk_count.append({
                    "Describtion": part["Describtion"],
                    "Part Number": part["Part Number"],
                    "Count": count
                })
    return lom_disk_count
#Extracts the data for fc ports
def lom_cards():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sysinfo_file = os.path.join(script_dir, "SystemOverallInfo", "SystemInfo.mylinux")

    with open(sysinfo_file, "r") as file:
        file_content = file.read()

    card_types = [
        {
            "section": "FC Cards",
            "label": "FC",
            "pattern": r"(\d+):\d+\.\d+ Fibre Channel:\s+(.+)"
        },
        {
            "section": "Network Cards",
            "label": "NIC",
            "pattern": r"(\d+):\d+\.\d+ Ethernet controller:\s+(.+)"
        }
    ]

    result = []

    for card_type in card_types:
        section = card_type["section"]
        label = card_type["label"]
        pattern = re.compile(card_type["pattern"])

        in_section = False
        grouped = defaultdict(list)

        for line in file_content.splitlines():
            line = line.strip()
            if line == section:
                in_section = True
                continue
            elif in_section and line.startswith("-="):
                break
            elif in_section:
                match = pattern.match(line)
                if match:
                    prefix = match.group(1)
                    model = match.group(2)
                    grouped[(prefix, model)].append(line)

        for (prefix, model), lines in grouped.items():
            result.append({
                "port count": len(lines),
                "Card": label,
                "model": model
            })

    return result
#Counts the FC/ISCSI ports from the output of lomfc_port
def merge_duplicate_dicts(dict_list):
    # Convert each dict into a hashable sorted tuple
    hashed = [tuple(sorted(d.items())) for d in dict_list]

    counts = Counter(hashed)

    # Reconstruct merged dicts with 'count' added
    result = []
    for items, count in counts.items():
        merged_dict = dict(items)
        merged_dict["count"] = count
        result.append(merged_dict)
    return result
#This part parses the information genereted by lom_ports into format suitable for LOM sheet
def lom_card_parcer(cards):
    lom_cards = []
    for card in cards:
        speed = 0
        Type = card["Card"]
        ports = card["port count"]
        count = card["count"]
        if "2261" in card.get("model", ""):
            speed = "16Gb"
        if "2071" in card.get("model", ""):
            speed = "16Gb"
        if "10GbE" in card.get("model", ""):
            speed = "10Gb"
        if "10-Gigabit" in card.get("model", ""):
            speed = "10Gb"
        elif "Gigabit" in card.get("model", ""):
            speed = "1Gb"
        if "8Gb" in card.get("model", ""):
            speed = "8Gb"
        if "16Gb" in card.get("model", ""):
            speed = "16Gb"
        for part in partnums_card:
            if ports != 1 and part["Type"] == Type and part["Ports"] == ports and part["Speed"] == speed:
                lom_cards.append({
                    "Describtion": part["Describtion"],
                    "Part Number": part["Part Number"],
                    "Count": count
                })
    return lom_cards
def is_valid_json_file(file_path):
    try:
        with open(file_path, 'r') as f:
            json.load(f)
        return True
    except (ValueError, json.JSONDecodeError) as e:
        #print(f"Invalid JSON: {e}")
        return False
    except FileNotFoundError:
        #print("File not found")
        return False
#Gather data for pool sheet
def pool_info():
    #it should be in the format of [Name, RAID type, number of disks, Size, LUNS, Hotspares]
    pool_data = []
    #Read DG,VD, Name, RAID type, size
    script_dir = os.path.dirname(os.path.abspath(__file__))

    storcli_Dall_path = os.path.join(script_dir, "SystemOverallInfo", "storcli-Dall-show-all.mylinux")
    lvm_info_path = os.path.join(script_dir, "SystemOverallInfo", "LVMInfo.mylinux")
    storcli_vall_path = os.path.join(script_dir, "SystemOverallInfo", "storcli-Vall-show.mylinux")
    pattern = r"""
        ^\s*                    # Start of line, optional whitespace
        (\d+)\s*\/\s*(\d+)      # DG/VD (flexible spacing around /)
        \s+(RAID(?:[0-9]+|10|50)) # RAID type (now includes RAID10 explicitly)
        \s+\S+\s+\S+\s+\S+      # Skip 3 columns (State, Access, Consist)
        \s+\S+                  # Cache type (any non-whitespace)
        \s+-\s+                 # sCC column
        ([\d.,]+\s+[TGM]B)      # Size (with optional , or .)
        \s+(\S+)                # Name (no whitespace)
        \s*$                    # End of line
    """
    #We need to check whether vall show is in json format
    if is_valid_json_file(storcli_vall_path):
        with open(storcli_vall_path, "r") as file:
            data = json.load(file)
            for controller in data.get("Controllers", []):
                virtual_drives = controller.get("Response Data", {}).get("Virtual Drives", [])
                for drive in virtual_drives:
                # Split DG/VD into separate components
                    dg_vd = drive.get("DG/VD", "").split('/')
                    if len(dg_vd) != 2:
                        continue                    
                    dg, vd = dg_vd
                    pool_data.append({
                        'dg': dg.strip(),
                        'vd': vd.strip(),
                        'Pool Name': drive.get("Name", ""),
                        'RAID Type': drive.get("TYPE", ""),
                        'Pool Size': drive.get("Size", ""),
                    })
    else:      
        with open(storcli_vall_path, "r") as file:
            vall_content = file.read()
            regex = re.compile(pattern, re.VERBOSE | re.MULTILINE)
            for match in regex.finditer(vall_content):
                dg, vd, raid_type, size, name = match.groups()
                pool_data.append({
                    "dg" : dg,
                    "vd" : vd,
                    "Pool Name" : name,
                    "Pool Size" : size,          
                    "RAID Type" : raid_type,
                })
    #Finding drive numbers and hotspares for each pool
    with open(storcli_Dall_path, "r") as file:
        dall_content = file.read()
        dg_section_match = re.search(
            r'DG Drive LIST :\s*\n=+\s*\n(.*?)\nTotal Drive Count',
            dall_content,
            re.DOTALL
        )
        dg_counts = defaultdict(int)
        if not dg_section_match:
           print("DG Drive LIST section not found in file")
           return
    
        dg_section = dg_section_match.group(1)
      
        pattern = r'^\s*\d+:\d+\s+\d+\s+\w+\s+(\d+)\s+([\d.]+\s*(?:TB|GB))\s+\w+\s+\w+\s+\w+\s+\w+\s+\w+\s+\w+.*$'  
        for line in dg_section.split('\n'):
            if line.strip():  # Skip empty lines
                match = re.match(pattern, line)
                if match:
                    dg_number = match.group(1)
                    drive_size = match.group(2)
                    dg_counts[dg_number] += 1
        for raid in pool_data:
            dg = str(raid["dg"])  # Ensure we compare strings to strings
            if dg in dg_counts:
                raid["Number of disks"] = dg_counts[dg]  # Add count to existing dictionary
                raid["Drive Size"] = drive_size

       #Catch Hotspare enclosure and slot
    # Extract the TOPOLOGY section
        section_match = re.search(
            r'TOPOLOGY :\s*\n=+\s*\n(.*?)\nVD LIST :\s*\n=+',
            dall_content,
            re.DOTALL
        )
        if not section_match:
            print("TOPOLOGY section not found in file")
            return []
        topology_content = section_match.group(1)
        dhs_drives = []
        pattern = r'^\s*(\d+)\s+-\s+-\s+(\d+:\d+)\s+\d+\s+DRIVE\s+DHS\s+.*$'
    
        for line in topology_content.split('\n'):
            line = line.strip()
            match = re.match(pattern, line)
            if match:
                dg_number = match.group(1)
                eid_slot = match.group(2)
                dhs_drives.append({
                    'DG': dg_number,
                    'EID:Slot': eid_slot
                })
        #Add Hotspare info
# First ensure all raid dictionaries have the 'Hotspares' key initialized
        for raid in pool_data:
            raid['Hotspares'] = ''
        for raids in pool_data:     
            for dhs in dhs_drives:
                if raids["dg"] == dhs["DG"]:
                    if raids["Hotspares"]:
                        raids["Hotspares"] += f", {dhs['EID:Slot']}"
                    else:
                        raids["Hotspares"] = dhs['EID:Slot']
        #fill empty hotspare cells with -
        for raid in pool_data:
            if raid["Hotspares"] == '':
                raid["Hotspares"] = "-"
    #Adding information for luns of each pool. Names plus size. (for now we merge the lun name and size with this format )
    with open(lvm_info_path, "r") as file:
         # Split the log into blocks for each disk.
        content = file.read()
        lv_section = re.search(
            r"lvdisplay(.*?)-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=",
            content,
            re.DOTALL
        )
        if lv_section:       
            disk_blocks = re.findall(
                r"--- Logical volume ---(.*?)Block device\s+",
                lv_section.group(1), 
                re.DOTALL
            )
            disk_blocks = [block.strip() for block in disk_blocks]
            pool_data_fn = []
        for block in disk_blocks:
            for raids in pool_data:
                lun_count= 0
                pool_name_match = re.search(r"VG Name\s+(\S+)", block, re.IGNORECASE)
                pool_name = pool_name_match.group(1)
                if pool_name == raids["Pool Name"]:
                    lun_name_match = re.search(r"LV Name\s+(\S+)", block, re.IGNORECASE)
                    lun_size_match = re.search(r"LV Size\s+(\S+\s+(\S))", block, re.IGNORECASE)
                    lun_name = lun_name_match.group(1)
                    lun_size = lun_size_match.group(1)
                    raid_copy = raids.copy()
                    raid_copy["LUN Name"] = lun_name_match.group(1)
                    raid_copy["LUN Size"] = lun_size_match.group(1)
                    pool_data_fn.append(raid_copy)
    # Define the desired field order
    field_order = ['dg', 'vd', 'Pool Size', 'Drive Size', 'Pool Name', 'LUN Name', 'LUN Size', 'Number of disks', 'Hotspares', 'RAID Type']
    # Create new sorted dictionaries
    sorted_pool_data = []
    for item in pool_data_fn:
        sorted_item = {field: item[field] for field in field_order if field in item}
        sorted_pool_data.append(sorted_item)
    
    #Sort the results according to dg number
    sorted_blocks = sorted(sorted_pool_data, key=lambda x: int(x['dg']))

    return sorted_blocks
    
#Extract the logs
def extractor():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    rust_binary = os.path.join(script_dir, "extractor.exe" if os.name == "nt" else "extractor")
    
    if not os.path.isfile(rust_binary):
        raise FileNotFoundError(f"Extractor binary not found at {rust_binary}")

    try:
        # Launch Rust binary (non-blocking)
        subprocess.Popen([rust_binary])
        print("Extractor started in background.")
        
        # Simulate work with extracted files
        print("Processing extracted files...")
        time.sleep(5)  # Replace with actual work
        while True:
            confirm = os.path.join(script_dir, "extracted_confirm")
            if os.path.isfile(confirm):
                print("Extraction Confirmed")
                response = input("Delete extracted files after finishing? (y/n): ").lower()
                if response in ['y', 'n']:
                    break
                print("Please enter 'y' or 'n'")
        
        # Create appropriate signal file
        signal_file = "delete_confirmed.flag" if response == 'y' else "delete_cancelled.flag"
        Path(signal_file).touch()

        
    except subprocess.CalledProcessError as e:
        print(f"[Python] Rust extractor failed: {e}")
        return False
# Reorder columns to make "Enclosure/Slot" the first column
def reorder_columns(data):
    return [{"En/Slot": disk.get("En/Slot", "N/A"), **disk} for disk in data]
def get_date():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sysinfo_file = os.path.join(script_dir, "SystemOverallInfo", "SystemInfo.mylinux")
    log_files = glob.glob(os.path.join(script_dir, "full_log*"))
    log_file = log_files[0]
    try:
        with open(sysinfo_file, 'r') as file:
            content = file.read()

            # Search for the time pattern
            match = re.search(r'Time\s+(\d{4}/\d{2}/\d{2})', content)

            if match:
                original_date = match.group(1)
                return original_date
            else:
                #return time of system
                timestamp = os.path.getmtime(log_file)
                modified_date = datetime.fromtimestamp(timestamp)
                formatted_date = modified_date.strftime("%Y/%m/%d")
                return formatted_date

    except FileNotFoundError:
        print(f"Error: File {file_path} not found")
        return None
    except Exception as e:
        print(f"Error processing file: {e}")
        return None
def parse_date(date_str):
    """Convert 'YYYY/MM/DD' string into year, month, day integers"""
    year, month, day = map(int, date_str.split('/'))
    return year, month, day
#Convert date from Gregorian to Jalali
def convertdate(year, month, day):
    result =  dict()
    array = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
    if year <= 1600:
        year -= 621
        result["year"] = 0
    else :
        year -= 1600
        result["year"] = 979
    temp = year + 1 if year > 2 else year
    days = (int((int((temp + 3) / 4)))) + (365 * year) - (int((int((temp + 99) / 100)))) - 80 + array[month - 1] + (int((int((temp + 399) / 400)))) + day
    result["year"] = result.get("year") + 33 * (int((int(days / 12053))))
    days %= 12053
    result["year"] = result.get("year") + 4 * (int((int(days / 1461))))
    days %= 1461
    if days > 365:
        result["year"] = result.get("year") + int((int((days - 1) / 365)))
        days = (days - 1) % 365
    result["month"] = 1 + int((int(days / 31))) if (days < 186) else 7 + int((int((days - 186) / 30)))
    result["day"] = 1 + ((days % 31) if (days < 186) else ((days - 186) % 30))
    jalalidate = str(result.get("year")) + "_" + str(result.get("month")) + "_" + str(result.get("day"))
    return jalalidate
def get_ID():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    log_files = [
        f for f in glob.glob(os.path.join(script_dir, "full_log*"))
        if os.path.isfile(f)  # Key change: filter out directories
    ]
    log_file = log_files[0]
    filename = os.path.basename(log_files[0])
    match = re.search(r'full_log(.+?)_(\d{4}-\d{2}-\d{2})', filename)
    if match:
        ID = match.group(1)
        return ID
    else:
        return " "

def output_name(ID, Date):
    return "smart" + "-"+ ID + "_" + Date + ".xlsx"

#Extract files
script_dir = os.path.dirname(os.path.abspath(__file__))
if not os.path.isfile(os.path.join(script_dir, 'version')):
    extractor()
# Path to the smarts.mylinux file in the /SystemOverallInfo directory
smarts_file_path = os.path.join(script_dir, "SystemOverallInfo", "smarts.mylinux")
target_data = extract_host_info()

# Path to the storcli-Sall-show-all.mylinux file in the /SystemOverallInfo directory
storcli_file_path = os.path.join(script_dir, "SystemOverallInfo", "storcli-Sall-show-all.mylinux")
#Setting the output name
log_date = get_date()
year, month, day = parse_date(log_date)
jalalidate = convertdate(year, month, day)
ID = get_ID()

# Read the log files
try:
    with open(smarts_file_path, "r") as file:
        smarts_content = file.read()
    with open(storcli_file_path, "r") as file:
        storcli_content = file.read()
except FileNotFoundError:
    print(f"Error: The required files were not found in the /SystemOverallInfo directory.")
    exit(1)

# Extract SSD, HDD, and device info data
ssd_data = extract_ssd_parameters(smarts_content)
# Check if OS disks are present
hdd_data = extract_hdd_parameters(smarts_content)
ssd_lom = []
hdd_lom = []
if ssd_data:
    ssd_lom = lom_disk(ssd_data)
if hdd_data:
    hdd_lom = lom_disk(hdd_data)
# Extract host information
host_data = extract_host_info()
# Extract General Device Info
sys_info = extract_sysinfo()
# Extract enclosure/slot information
serial_numbers = set([disk["Serial Number"] for disk in ssd_data + hdd_data if disk["Serial Number"] != "Unknown"])
enclosure_slot_data = extract_enclosure_slot_info(storcli_content, serial_numbers)
#lom for cards
lom_cards_results = lom_cards()
if lom_cards_results:
   lom_cards_fixed = merge_duplicate_dicts(lom_cards_results)
   lom_cards_final = lom_card_parcer(lom_cards_fixed)
# Add enclosure/slot information to SSD and HDD data
for disk in ssd_data + hdd_data:
    if disk["Serial Number"] in enclosure_slot_data:
        disk["Disk State"] = enclosure_slot_data[disk["Serial Number"]]["Disk State"]
        disk["Media Err"] = enclosure_slot_data[disk["Serial Number"]]["media_error_count"]
        disk["Other Err"] = enclosure_slot_data[disk["Serial Number"]]["other_error_count"]        
        disk["Shield Cnt"] = enclosure_slot_data[disk["Serial Number"]]["shield_counter"]
        disk["Predictive Failure"] = enclosure_slot_data[disk["Serial Number"]]["predictive_failure_count"]
        disk["En/Slot"] = enclosure_slot_data[disk["Serial Number"]]["enclosure_slot"]

    else:
        disk["En/Slot"] = ""
is_os = any("250 GB" in str(value) for d in ssd_data for value in d.values())
#OS Disk not found therefore we extract from pmc output
if not is_os:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pmc_files = glob.glob(os.path.join(script_dir, "output.txt"))
    if pmc_files:
        with open(pmc_files[0], "r") as f:
            log = f.read()
        ssd_os = extract_ssd_parameters(log)
        for disk in ssd_os:
            disk["En/Slot"] = "Not in RC"
        ssd_data.extend(ssd_os)
#Merge both ssd and hdd data for SMART
diskdata = hdd_data + ssd_data
device_data = extract_device_info(smarts_content, diskdata)

def write_slot_info_sheet(writer, slot_data):
    wb = writer.book
    ws = wb.create_sheet("Slot Info")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")     # Light grey
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Light green for FC
    blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")     # Light blue for iSCSI

    slots_order = [str(i) for i in range(6, 0, -1)]  # Slots 6 to 1
    attributes = ["Port", "WWN", "Connection Type", "Speed"]
    cols_per_slot = len(attributes)

    # Prepare a dict for quick access
    slot_map = {slot['slot']: slot for slot in slot_data}

    # Write first row: slot header merged across cols_per_slot columns
    for i, slot_num in enumerate(slots_order):
        start_col = i * cols_per_slot + 1
        end_col = start_col + cols_per_slot - 1
        cell = ws.cell(row=1, column=start_col)
        cell.value = f"Slot{slot_num}"
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply light grey fill if the slot has data
        slot = slot_map.get(f"Slot{slot_num}")
        if slot and slot['total_ports'] > 0:
            for col in range(start_col, end_col + 1):
                ws.cell(row=1, column=col).fill = grey_fill
    # Write second row: port type merged across same columns, empty if missing slot
    for i, slot_num in enumerate(slots_order):
        slot_key = f"Slot{slot_num}"
        slot = slot_map.get(slot_key)
        start_col = i * cols_per_slot + 1
        end_col = start_col + cols_per_slot - 1
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = ws.cell(row=2, column=start_col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if slot is None or not slot.get('port_type'):
            cell.value = ""
        else:
            cell.value = slot['port_type'].upper()
            # Fill color by port type
            fill = green_fill if slot['port_type'].lower() == "fc" else blue_fill
            for col in range(start_col, end_col + 1):
                ws.cell(row=2, column=col).fill = fill        
    # Write third row: attribute headers
    for i, slot_num in enumerate(slots_order):
        slot_key = f"Slot{slot_num}"
        slot = slot_map.get(slot_key)
        start_col = i * cols_per_slot + 1

        if slot is None or slot['total_ports'] == 0:
            for j in range(cols_per_slot):
                ws.cell(row=3, column=start_col + j).value = ""
        else:
            for j, attr in enumerate(attributes):
                label = attr
                if attr == "Connection Type" and slot and slot.get('port_type', '').lower() == 'iscsi':
                    label = "MAC Address"
                ws.cell(row=3, column=start_col + j).value = label    # Determine max number of rows needed
    max_ports = max((slot['total_ports'] for slot in slot_map.values()), default=0)

    # Write port data (row 4 onward)
    for row_idx in range(max_ports):
        for i, slot_num in enumerate(slots_order):
            slot_key = f"Slot{slot_num}"
            slot = slot_map.get(slot_key)
            start_col = i * cols_per_slot + 1

            if slot is None or row_idx >= slot['total_ports']:
                for offset in range(cols_per_slot):
                    ws.cell(row=4 + row_idx, column=start_col + offset).value = ""
                continue

            port_info = slot['ports'].get(row_idx, {})
            port = row_idx
            wwn = port_info.get('wwn', '')

            # Use MAC address instead of connection type for iSCSI
            if slot.get('port_type', '').lower() == 'iscsi':
                connection = port_info.get('mac_address') or port_info.get('mac') or "-"
            else:
                connection = port_info.get('connection_type', '-')  # FC fallback

            speed = port_info.get('speed', '-')
            if not speed or speed.lower() == "unknown":
                speed = "-"

            # Write to sheet
            ws.cell(row=4 + row_idx, column=start_col).value = port
            ws.cell(row=4 + row_idx, column=start_col + 1).value = wwn
            ws.cell(row=4 + row_idx, column=start_col + 2).value = connection
            ws.cell(row=4 + row_idx, column=start_col + 3).value = speed

    # Adjust column widths
    for col in range(1, cols_per_slot * len(slots_order) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


ssd_data = reorder_columns(ssd_data)
hdd_data = reorder_columns(hdd_data)

# Remove rows where "Enclosure/Slot" is "N/A"
ssd_data = [disk for disk in ssd_data]
hdd_data = [disk for disk in hdd_data]
#Create output name
output = output_name(ID, jalalidate)
# Create an Excel writer
excel_path = output
#Create slot info
slot_info = extract_slot_port_info()

pool_data = pool_info()
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Write SMART data to first sheet
    df_smart = pd.DataFrame(ssd_data + hdd_data)

    # Ensure empty rows are preserved in the DataFrame
    # Replace empty strings with NaN (optional, but helps with consistency)
    df_smart.replace("", pd.NA, inplace=True)

    # Write the DataFrame to the Excel file
    df_smart.to_excel(writer, sheet_name="SMART Data", index=False)

    # Write device temperature & powered-up hours to the second sheet (only if non-empty)
    if device_data:
        df_device = pd.DataFrame(device_data)
        df_device.to_excel(writer, sheet_name="Device Info", index=False)
    if pool_data:
        df_host = pd.DataFrame(pool_data)
        df_host.to_excel(writer, sheet_name="Pool Data", index=False)
    # Write host information to the third sheet (only if non-empty)
    if host_data:
        df_host = pd.DataFrame(host_data)
        df_host.to_excel(writer, sheet_name="Host Info", index=False)
    if sys_info:
        df_host = pd.DataFrame(sys_info)
        df_host.to_excel(writer, sheet_name="General System Info", index=False)
    if slot_info:
        write_slot_info_sheet(writer, slot_info)
    if ssd_lom or hdd_lom or lom_cards_final:
        dfs = []
        if ssd_lom:
            dfs.append(pd.DataFrame(ssd_lom))
        if hdd_lom:
            dfs.append(pd.DataFrame(hdd_lom))
        if lom_cards_final:
            dfs.append(pd.DataFrame(lom_cards_final))
        df_combined = pd.concat(dfs, ignore_index=True)
        df_combined.to_excel(writer, sheet_name="LOM", index=False)
                        
# Open the Excel file and format it
wb = load_workbook(excel_path)
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="de0a0a", end_color="FFFF00", fill_type="solid")
#Colour cells based on thresholds
if "SMART Data" in wb.sheetnames:
    ws = wb["SMART Data"]
    for row in ws.iter_rows(min_row=2):  # Skip header row (row 1)
        # Extract relevant cells
        threshold_cell = row[7]  # Column H (Threshold)
        value_cell = row[8]      # Column I (Value)
        raw_value_cell = row[9]  # Column J (Raw Value)
        param_value_cell = row[6]

        param_str = param_value_cell.value
        threshold_str = threshold_cell.value
        value_str = value_cell.value
        raw_value_str = raw_value_cell.value
        #We dont want to color this cell
        is_unused_rsvd = False
        if param_str == "Unused_Rsvd_Blk_Cnt_Tot":
            is_unused_rsvd = True
        if threshold_str:
            threshold_value = float(threshold_str)
        # Skip rows with missing/invalid thresholds
        if not threshold_str or threshold_str == "-":
            continue

        try:
            threshold_caution = float(threshold_str) * 1.5
            threshold_warning = float(threshold_str) * 1.2
        except (ValueError, TypeError):
            continue  # Skip non-numeric thresholds

        # Check Value or Raw Value
        compare_value = None
        if value_str not in (None, "-", ""):
            try:
                compare_value = float(value_str)
                #To avoid colouring cells when value is 100
                if compare_value == 100:
                    compare_value = 200

            except (ValueError, TypeError):
                pass

        if compare_value is None:  # Fallback to Raw Value
            if raw_value_str not in (None, "-", ""):
                try:
                    #because here raw value is smaller than thresold in normal state
                    threshold_caution = float(threshold_str) * -0.8
                    threshold_warning = float(threshold_str) * -0.9
                    compare_value = float(raw_value_str) * -1
                    threshold_value = threshold_value * -1
                   # print(threshold_warning)
                except (ValueError, TypeError):
                    continue  # Skip invalid values
            else:
                continue

        # Highlight if value < threshold
        if compare_value <= threshold_value:
            raw_value_cell.fill = red_fill
        elif compare_value <= threshold_warning:
            raw_value_cell.fill = orange_fill
        elif compare_value <= threshold_caution:
            raw_value_cell.fill = yellow_fill
        if is_unused_rsvd:
            raw_value_cell.fill = PatternFill()

# Function to merge cells for a specific column
def merge_cells_for_column(ws, col_idx):
    prev_value = None
    start_row = 2  # Start from the second row (first row is headers)

    for row in range(2, ws.max_row + 1):
        current_value = ws.cell(row=row, column=col_idx).value
        if current_value == prev_value:
            continue
        if prev_value is not None and start_row is not None:
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=row-1, end_column=col_idx)
            ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")
        prev_value = current_value
        start_row = row

    if prev_value is not None and start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")

# Function to adjust column widths automatically
def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 4
deep_blue_fill = PatternFill(start_color="b8cbdf", end_color="b8cbdf", fill_type="solid")
# Format all sheets except "Device Info"
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    #Color the first row of sheets
    if sheet_name != "Slot Info":
        for cell in ws[1]:# First row
            cell.fill = deep_blue_fill
    if sheet_name not in ("Device Info", "Slot Info", "LOM", "Pool Data"):# Skip merging for "Device Info" sheet
        for column in range(1,7):
            merge_cells_for_column(ws, column)
        for column in range(11,16):
            merge_cells_for_column(ws, column)
    adjust_column_widths(ws)  # Adjust column widths for all sheets
if "Host Info" in wb.sheetnames: 
    host_info_sheet = wb["Host Info"]
    merge_cells_for_column(host_info_sheet, 4)  # Merge "Initiators" column (column 4)
    merge_cells_for_column(host_info_sheet, 5)  # Merge "Targets" column (column 5) 
    merge_cells_for_column(host_info_sheet, 6)  # Merge "Connection type" column (column 6) 
if "Pool Data" in wb.sheetnames: 
    pooldata = wb["Pool Data"]
    merge_cells_for_column(pooldata, 1)  
    merge_cells_for_column(pooldata, 2)  
    merge_cells_for_column(pooldata, 3)  
    merge_cells_for_column(pooldata, 4)  
    merge_cells_for_column(pooldata, 5)  
    merge_cells_for_column(pooldata, 7)
    merge_cells_for_column(pooldata, 8) 
    merge_cells_for_column(pooldata, 9)  
    merge_cells_for_column(pooldata, 6)
    merge_cells_for_column(pooldata, 10)  
wb.save(excel_path)
print("SMART data, device info, and host info extracted and written to the Excel file with proper formatting.")
Path("python_done.flag").touch()
