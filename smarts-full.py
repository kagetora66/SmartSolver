import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import csv
import os
import glob
import sqlite3

# Define required parameters
ssd_params = [
    "Reallocated_Sector_Ct",
    "Power_On_Hours",
    "Wear_Leveling_Count",
    "Used_Rsvd_Blk_Cnt_Tot",
    "Runtime_Bad_Block",
    "Reported_Uncorrect",
    "Hardware_ECC_Recovered",
    "Total_LBAs_Written"
]

hdd_params = [
    "Elements in grown defect list",
    "Total Uncorrected Errors",
    "Accumulated start-stop cycles",
    "Accumulated load-unload cycles"
]

# Regular expression to extract SMART attributes
smart_pattern = re.compile(
    r"(\d+)\s+([\w_]+)\s+0x[0-9a-fA-F]+\s+(\d+)\s+\d+\s+\d+\s+\w+-?\w*\s+\w+\s+-\s+(\d+)"
)
# Function to extract SSD parameters
def extract_ssd_parameters(log_content):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    
    for block in disk_blocks:
        is_ssd = not re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        is_sas_ssd = re.search(r"Transport protocol:\s+SAS", block, re.IGNORECASE)
        
        if is_ssd:
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            
            if is_sas_ssd:
                # Extract HDD parameters for SAS SSDs
                elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
                start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
                load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)

                # Extract Total Uncorrected Errors
                read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
                write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
                
                read_error_value = int(read_error_match.group(1)) if read_error_match else 0
                write_error_value = int(write_error_match.group(1)) if write_error_match else 0
                total_uncorrected_errors = read_error_value + write_error_value

                # Sort the parameters in the desired order
                hdd_values = [
                    ("Elements in grown defect list", elements_grown_defect),
                    ("Total Uncorrected Errors", total_uncorrected_errors),
                    ("Accumulated start-stop cycles", start_stop_cycles),
                    ("Accumulated load-unload cycles", load_unload_cycles)
                ]
                
                for param, match in hdd_values:
                    if isinstance(match, int):  # For Total Uncorrected Errors
                        raw_value = str(match)
                        data.append({
                            "Serial Number": serial_number,
                            "Parameter": param,
                            "Value": "-",
                            "Raw Value": raw_value
                        })
                    elif match:
                        raw_value = match.group(1) if hasattr(match, "group") else str(match)
                        data.append({
                            "Serial Number": serial_number,
                            "Parameter": param,
                            "Value": "-",
                            "Raw Value": raw_value
                        })  
            else:
                # Extract SSD parameters for SATA SSDs
                smart_matches = smart_pattern.findall(block)
                for match in smart_matches:
                    attr_id, attr_name, value, raw_value = match
                    if attr_name in ssd_params:
                        data.append({
                            "Serial Number": serial_number,
                            "Parameter": attr_name,
                            "Value": value,
                            "Raw Value": raw_value
                        })
            
            # Add an empty row after each disk's data
            data.append({
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
    return data

def extract_hdd_parameters(log_content):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    
    for block in disk_blocks:
        is_hdd = re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        if is_hdd:
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            
            elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
            start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
            load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)

            # Extract Total Uncorrected Errors
            read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
            write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
            
            read_error_value = int(read_error_match.group(1)) if read_error_match else 0
            write_error_value = int(write_error_match.group(1)) if write_error_match else 0
            total_uncorrected_errors = read_error_value + write_error_value

            # Sort the parameters in the desired order
            hdd_values = [
                ("Elements in grown defect list", elements_grown_defect),
                ("Total Uncorrected Errors", total_uncorrected_errors),
                ("Accumulated start-stop cycles", start_stop_cycles),
                ("Accumulated load-unload cycles", load_unload_cycles)
            ]
            
            for param, match in hdd_values:
                if isinstance(match, int):  # For Total Uncorrected Errors
                    raw_value = str(match)
                    data.append({
                        "Serial Number": serial_number,
                        "Parameter": param,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
                elif match:
                    raw_value = match.group(1) if hasattr(match, "group") else str(match)
                    data.append({
                        "Serial Number": serial_number,
                        "Parameter": param,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
            # Add an empty row after each disk's data
            data.append({
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
    return data

# Function to extract HDD device info (only if both values are found)
def extract_device_info(log_content):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)

    for block in disk_blocks:
        is_hdd = re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        if is_hdd:
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else None

            temp_match = re.search(r"Current Drive Temperature:\s+(\d+)", block)
            hours_match = re.search(r"number of hours powered up\s+=\s+([\d.]+)", block)

            if serial_number and temp_match and hours_match:
                temperature = f"{temp_match.group(1)}"
                hours = hours_match.group(1)
                data.append({
                    "Device": serial_number,
                    "Temperature": temperature,
                    "Powered Up Hours": hours
                })

    return data
def extract_enclosure_slot_info(log_content, serial_numbers):
    enclosure_slot_data = {}
    
    # Split the log content into lines
    lines = log_content.splitlines()
    
    # Iterate through the lines to find serial numbers and their corresponding Drive lines
    for i, line in enumerate(lines):
        # Check if the line contains a serial number
        if line.strip().startswith("SN ="):
            serial = line.split("=")[1].strip()
            if serial in serial_numbers:
                # Search backward for the Drive line
                for j in range(i, max(i - 20, -1), -1):  # Look back up to 20 lines
                    if lines[j].strip().startswith("Drive /c"):
                        # Extract enclosure and slot from the Drive line
                        drive_line = lines[j].strip()
                        enclosure = drive_line.split("/e")[1].split("/")[0]
                        slot = drive_line.split("/s")[1].split()[0]
                        enclosure_slot_data[serial] = f"{enclosure}/{slot}"
                        break
    
    return enclosure_slot_data
# Function to extract host information from SCST configuration
def extract_host_info():
    # Directories
    scst_dir = "./SCST"
    db_dir = "./Database"

    # Find the most recent SCST configuration file
    scst_files = sorted(glob.glob(os.path.join(scst_dir, "scst_*.conf")), reverse=True)
    if not scst_files:
        print("Error: No 'scst_*.conf' files found in /SCST directory.")
        exit(1)

    input_file = scst_files[0]  # Select the most recent file
    output_file = "access_points.csv"

    try:
        # Read the input file
        with open(input_file, "r") as file:
            input_text = file.read()

        # Regular expression to match GROUP sections
        group_pattern = r"GROUP\s+([\w-]+)\s*\{([\s\S]*?)\}"  # Matches GROUP name and its content
        lun_pattern = r"LUN\s+(\d+)\s+([\w_]+)"  # Matches LUN numbers and names
        initiator_pattern = r"INITIATOR\s+([0-9a-fA-F:]+)"  # Matches INITIATOR addresses

        # Extract all groups
        groups = re.findall(group_pattern, input_text)

        # Prepare data for Excel
        host_data = []

        # Process each group
        for group_name, group_content in groups:
            # Extract LUNs and INITIATORS specific to this group
            luns = re.findall(lun_pattern, group_content)
            initiators = re.findall(initiator_pattern, group_content)

            # Remove duplicates for LUNs and Initiators
            unique_luns = sorted(set(luns), key=lambda x: int(x[0]))  # Sort by LUN number
            unique_initiators = sorted(set(initiators))

            # Database processing starts here
            sab_db_file = os.path.join(db_dir, "sab.db")  # SQLite database file

            # Connect to the SQLite database
            conn = sqlite3.connect(sab_db_file)
            cursor = conn.cursor()

            # Query to find the `to_host_id` for each initiator in the "hostinitiators" table
            hostinitiators_query = """
            SELECT initiator_name, to_host_id 
            FROM hostinitiators
            WHERE initiator_name = ?
            """

            # Query to find the `name` in the "host" table using `to_host_id`
            host_query = """
            SELECT NAME 
            FROM host
            WHERE ID = ?
            """

            # Map initiators to their host names
            host_map = {}
            for initiator in unique_initiators:
                cursor.execute(hostinitiators_query, (initiator,))
                to_host_row = cursor.fetchone()

                if to_host_row:
                    to_host_id = to_host_row[1]
                    cursor.execute(host_query, (to_host_id,))
                    host_row = cursor.fetchone()

                    if host_row:
                        host_map[initiator] = host_row[0]
                    else:
                        host_map[initiator] = "Not Found"
                else:
                    host_map[initiator] = "Not Found"

            # Group LUNs and Initiators by host
            host_luns_initiators = {}
            for initiator, host_name in host_map.items():
                if host_name not in host_luns_initiators:
                    host_luns_initiators[host_name] = {"luns": set(), "initiators": set()}
                host_luns_initiators[host_name]["luns"].update([lun[1] for lun in unique_luns if lun[1] != "device_null"])  # Exclude device_null
                host_luns_initiators[host_name]["initiators"].add(initiator)

            # Prepare rows for each host, LUNs, and initiators
            for host_name, data in host_luns_initiators.items():
                # If there are no LUNs, still include the access point
                if not data["luns"]:
                    host_data.append({
                        "Access Point": group_name,
                        "Host": host_name,
                        "LUNs": "",  # Empty string for no LUNs
                        "Initiator Addresses": ", ".join(sorted(data["initiators"]))
                    })
                else:
                    # Include each LUN in a separate row
                    for lun in sorted(data["luns"]):
                        host_data.append({
                            "Access Point": group_name,
                            "Host": host_name,
                            "LUNs": lun,
                            "Initiator Addresses": ", ".join(sorted(data["initiators"]))
                        })

        return host_data

    except ValueError as e:
        print(f"Error: {e}")
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        if 'conn' in locals():
            conn.close()

# Get the directory of the script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Path to the smarts.mylinux file in the /SystemOverallInfo directory
smarts_file_path = os.path.join(script_dir, "SystemOverallInfo", "smarts.mylinux")

# Path to the storcli-Sall-show-all.mylinux file in the /SystemOverallInfo directory
storcli_file_path = os.path.join(script_dir, "SystemOverallInfo", "storcli-Sall-show-all.mylinux")

# Read the log files
try:
    with open(smarts_file_path, "r") as file:
        smarts_content = file.read()
    with open(storcli_file_path, "r") as file:
        storcli_content = file.read()
except FileNotFoundError:
    print(f"Error: The required files were not found in the /SystemOverallInfo directory.")
    exit(1)

# Extract SSD, HDD, and device info data
ssd_data = extract_ssd_parameters(smarts_content)
hdd_data = extract_hdd_parameters(smarts_content)
device_data = extract_device_info(smarts_content)

# Extract host information
host_data = extract_host_info()

# Extract enclosure/slot information
serial_numbers = set([disk["Serial Number"] for disk in ssd_data + hdd_data if disk["Serial Number"] != "Unknown"])
enclosure_slot_data = extract_enclosure_slot_info(storcli_content, serial_numbers)

# Add enclosure/slot information to SSD and HDD data
for disk in ssd_data + hdd_data:
    if disk["Serial Number"] in enclosure_slot_data:
        disk["Enclosure/Slot"] = enclosure_slot_data[disk["Serial Number"]]
    else:
        disk["Enclosure/Slot"] = "N/A"

# Reorder columns to make "Enclosure/Slot" the first column
def reorder_columns(data):
    return [{"Enclosure/Slot": disk.get("Enclosure/Slot", "N/A"), **disk} for disk in data]

ssd_data = reorder_columns(ssd_data)
hdd_data = reorder_columns(hdd_data)

# Remove rows where "Enclosure/Slot" is "N/A"
ssd_data = [disk for disk in ssd_data if disk["Enclosure/Slot"] != "N/A"]
hdd_data = [disk for disk in hdd_data if disk["Enclosure/Slot"] != "N/A"]

# Create an Excel writer
excel_path = 'smart_data.xlsx'
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Write SMART data to first sheet
    df_smart = pd.DataFrame(ssd_data + hdd_data)
    df_smart.to_excel(writer, sheet_name="SMART Data", index=False)

    # Write device temperature & powered-up hours to second sheet (only if non-empty)
    if device_data:
        df_device = pd.DataFrame(device_data)
        df_device.to_excel(writer, sheet_name="Device Info", index=False)

    # Write host information to third sheet (only if non-empty)
    if host_data:
        df_host = pd.DataFrame(host_data)
        df_host.to_excel(writer, sheet_name="Host Info", index=False)

# Open the Excel file and format it
wb = load_workbook(excel_path)

# Function to merge cells for a specific column
def merge_cells_for_column(ws, col_idx):
    prev_value = None
    start_row = 2  # Start from the second row (first row is headers)

    for row in range(2, ws.max_row + 1):
        current_value = ws.cell(row=row, column=col_idx).value
        if current_value == prev_value:
            continue
        if prev_value is not None and start_row is not None:
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=row-1, end_column=col_idx)
            ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")
        prev_value = current_value
        start_row = row

    if prev_value is not None and start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")

# Function to adjust column widths automatically
def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

# Format all sheets except "Device Info"
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if sheet_name != "Device Info":  # Skip merging for "Device Info" sheet
        merge_cells_for_column(ws, 1)  # Merge "Enclosure/Slot" column (column 1)
        merge_cells_for_column(ws, 2)  # Merge "Serial Number" column (column 2)
    adjust_column_widths(ws)  # Adjust column widths for all sheets
if "Host Info" in wb.sheetnames: 
    host_info_sheet = wb["Host Info"]
    merge_cells_for_column(host_info_sheet, 4)  # Merge "Enclosure/Slot" column (column 1)
wb.save(excel_path)

print("SMART data, device info, and host info extracted and written to smart_data.xlsx with proper formatting.")
