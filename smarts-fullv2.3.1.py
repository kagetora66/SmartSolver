import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import csv
import os
import glob
import sqlite3
import getpass
from zipfile import ZipFile
# Define required parameters
ssd_params = [
    "Reallocated_Sector_Ct",
    "Power_On_Hours",
    "Wear_Leveling_Count",
    "Used_Rsvd_Blk_Cnt_Tot",
    "Runtime_Bad_Block",
    "Reported_Uncorrect",
    "Hardware_ECC_Recovered",
    "Total_LBAs_Written",
    "Total Size Written (TB)"
]

hdd_params = [
    "Elements in grown defect list",
    "Total Uncorrected Errors",
    "Accumulated start-stop cycles",
    "Accumulated load-unload cycles"
]
micron_ssd_params = {
    "Raw_Read_Error_Rate": "1",
    "Reallocated_Sector_Ct": "5",
    "Reported_Uncorrect": "187",
    "Hardware_ECC_Recovered": "195",
    "Unused_Rsvd_Blk_Cnt_Tot": "180",
    "Total_LBAs_Written": "246"  # Unknown attribute at ID# 246
}
    
# Regular expression to extract SMART attributes
smart_pattern = re.compile(
    r"(\d+)\s+([\w_]+)\s+0x[0-9a-fA-F]+\s+(\d+)\s+\d+\s+\d+\s+\w+-?\w*\s+\w+\s+-\s+(\d+)"
)
# Function to extract SSD parameters
def extract_ssd_parameters(log_content):
    data = []
    # Split the log into blocks for each disk.
    disk_blocks = re.findall(
        r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)",
        log_content, re.DOTALL
    )
    
    for block in disk_blocks:
        # Determine disk type based on certain strings
        is_ssd = not re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        is_sas_ssd = re.search(r"Transport protocol:\s+SAS", block, re.IGNORECASE)
        is_micron_ssd = re.search(r"Device Model:\s+Micron", block, re.IGNORECASE)
        
        if is_ssd:
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            model_match =  re.search(r"Device Model:\s+(\S+)", block, re.IGNORECASE)
            device_model = serial_match.group(1) if model_match else "Unknown" 
            if is_sas_ssd:
                # --- SAS SSD extraction (existing logic) ---
                elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
                start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
                load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)
                read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
                write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
                
                read_error_value = int(read_error_match.group(1)) if read_error_match else 0
                write_error_value = int(write_error_match.group(1)) if write_error_match else 0
                total_uncorrected_errors = read_error_value + write_error_value

                hdd_values = [
                    ("Elements in grown defect list", elements_grown_defect),
                    ("Total Uncorrected Errors", total_uncorrected_errors),
                    ("Accumulated start-stop cycles", start_stop_cycles),
                    ("Accumulated load-unload cycles", load_unload_cycles)
                ]
                
                for param, match in hdd_values:
                    if isinstance(match, int):  # For Total Uncorrected Errors
                        raw_value = str(match)
                    elif match:
                        raw_value = match.group(1)
                    else:
                        continue
                    data.append({
                        "Serial Number": serial_number,
                        "Device Model": device_model, 
                        "Parameter": param,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
                    
            else:
                # --- Non-SAS SSD extraction ---
                # smart_pattern should be defined elsewhere to extract tuples:
                # (attr_id, attr_name, value, raw_value)
                smart_matches = smart_pattern.findall(block)
                total_lba_written = None  # To store Total_LBAs_Written when found
                
                if is_micron_ssd:
                    # Define the list of expected Micron SMART parameters.
                    micron_params = [
                        "Raw_Read_Error_Rate", 
                        "Reallocated_Sector_Ct", 
                        "Reported_Uncorrect", 
                        "Hardware_ECC_Recovered", 
                        "Unused_Rsvd_Blk_Cnt_Tot"
                    ]
                    
                    for match in smart_matches:
                        attr_id, attr_name, value, raw_value = match
                        # Check for the known Micron parameters.
                        if attr_name in micron_params:
                            data.append({
                                "Serial Number": serial_number,
                                "Device Model": device_model,
                                "Parameter": attr_name,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        # Detect Total_LBAs_Written by its attribute ID "246"
                        elif attr_id.strip() == "246":
                            try:
                                total_lba_written = int(raw_value)
                            except ValueError:
                                total_lba_written = None
                            data.append({
                                "Serial Number": serial_number,
                                "Device Model": device_model,
                                "Parameter": "Total_LBAs_Written",
                                "Value": value,
                                "Raw Value": raw_value
                            })
                            
                    if total_lba_written is not None:
                        total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                        data.append({
                            "Serial Number": serial_number,
                            "Device Model": device_model,
                            "Parameter": "Total Size Written (TB)",
                            "Value": "-",
                            "Raw Value": f"{total_size_written_tb:.2f}"
                        })
                else:
                    # Existing logic for non-Micron SATA SSDs
                    for match in smart_matches:
                        attr_id, attr_name, value, raw_value = match
                        if attr_name in ssd_params:  # ssd_params defined elsewhere
                            if attr_name == "Total_LBAs_Written":
                                try:
                                    total_lba_written = int(raw_value)
                                except ValueError:
                                    total_lba_written = None
                            data.append({
                                "Serial Number": serial_number,
                                "Device Model": device_model,
                                "Parameter": attr_name,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                    if total_lba_written is not None:
                        total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                        data.append({
                            "Serial Number": serial_number,
                            "Device Model": device_model,
                            "Parameter": "Total Size Written (TB)",
                            "Value": "-",
                            "Raw Value": f"{total_size_written_tb:.2f}"
                        })
            
            # Add an empty row after each disk's data for readability.
            data.append({
                "Serial Number": "",
                "Device Model": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
            for i in range(len(data) - 1):
                if (
                    data[i]["Parameter"] == "Total_LBAs_Written" and
                    data[i + 1]["Parameter"] == "Unused_Rsvd_Blk_Cnt_Tot"
                ):
                    # Swap their positions
                    data[i], data[i + 1] = data[i + 1], data[i]
            
    return data

# Function to extract HDD parameters
def extract_hdd_parameters(log_content):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    
    for block in disk_blocks:
        is_hdd = re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        if is_hdd:
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"

            model_match =  re.search(r"Device Model:\s+(\S+)", block, re.IGNORECASE)
            model_match_hp = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)
            device_model = (
                    model_match.group(1) if model_match 
                    else model_match_hp.group(1) if model_match_hp
                    else "Unknown"
                    )
  
            elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
            start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
            load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)

            # Extract Total Uncorrected Errors
            read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
            write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
            
            read_error_value = int(read_error_match.group(1)) if read_error_match else 0
            write_error_value = int(write_error_match.group(1)) if write_error_match else 0
            total_uncorrected_errors = read_error_value + write_error_value

            # Sort the parameters in the desired order
            hdd_values = [
                ("Elements in grown defect list", elements_grown_defect),
                ("Total Uncorrected Errors", total_uncorrected_errors),
                ("Accumulated start-stop cycles", start_stop_cycles),
                ("Accumulated load-unload cycles", load_unload_cycles)
            ]
            
            for param, match in hdd_values:
                if isinstance(match, int):  # For Total Uncorrected Errors
                    raw_value = str(match)
                    data.append({
                        "Serial Number": serial_number,
                        "Device Model": device_model,
                        "Parameter": param,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
                elif match:
                    raw_value = match.group(1) if hasattr(match, "group") else str(match)
                    data.append({
                        "Serial Number": serial_number,
                        "Device Model": device_model,
                        "Parameter": param,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
            # Add an empty row after each disk's data
            data.append({
                "Serial Number": "",
                "Device Model": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
    return data

# Function to extract HDD device info (only if both values are found)
def extract_device_info(log_content):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    tmp_data = []
    for block in disk_blocks:
         serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
         serial_number = serial_match.group(1) if serial_match else None

         temp_match = re.search(r"Current Drive Temperature:\s+(\d+)", block)
         hours_match = re.search(r"number of hours powered up\s+=\s+([\d.]+)", block)

         if serial_number and temp_match:
             temperature = f"{temp_match.group(1)}"
             if hours_match:                
                 hours = hours_match.group(1)
                 tmp_data.append({
                     "Device": serial_number,
                     "Temperature": temperature,
                     "Powered Up Hours": hours
                 })
             else:
                data.append({
                     "Device": serial_number,
                     "Temperature": temperature,
                 })
    data.extend(tmp_data)
    return data

# Function to extract enclosure/slot information
def extract_enclosure_slot_info(log_content, serial_numbers):
    enclosure_slot_data = {}

    # Split the log content into lines
    lines = log_content.splitlines()

    current_shield_counter = None
    current_media_error_count = None
    current_other_error_count = None

    for i, line in enumerate(lines):
        line = line.strip()

        # Capture Shield Counter
        if line.startswith("Shield Counter"):
            current_shield_counter = line.split("=")[1].strip()

        # Capture Media Error Count
        elif line.startswith("Media Error Count"):
            current_media_error_count = line.split("=")[1].strip()

        # Capture Other Error Count
        elif line.startswith("Other Error Count"):
            current_other_error_count = line.split("=")[1].strip()

        # Detect Serial Number
        elif line.startswith("SN ="):
            serial = line.split("=")[1].strip()
            if serial in serial_numbers:
                # Search backward for Drive line
                for j in range(i, max(i - 20, -1), -1):
                    drive_line = lines[j].strip()
                    if drive_line.startswith("Drive /c"):
                        enclosure = drive_line.split("/e")[1].split("/")[0]
                        slot = drive_line.split("/s")[1].split()[0]

                        # Now assign all gathered info
                        enclosure_slot_data[serial] = {
                            "enclosure_slot": f"{enclosure}/{slot}",
                            "shield_counter": current_shield_counter,
                            "media_error_count": current_media_error_count,
                            "other_error_count": current_other_error_count
                        }
                        break

    return enclosure_slot_data
def extract_sysinfo():
    sysinfo_file = os.path.join(script_dir, "SystemOverallInfo", "SystemInfo.mylinux")
    version_file = os.path.join(script_dir, "version")
    sys_info = {}
    voltage_index = 1 #For separting the two power modules
    current_index = 1 #For separating the two power modules
    try:
        # Extract uptime and serial number from SystemInfo.mylinux
        with open(sysinfo_file, "r") as file:
            for line in file:
                uptime_match = re.search(r"up\s+(\d+)\s+days?", line)
                serial_match = re.search(r"Serial Number:\s*ZM(\S+)", line)
                voltage_match = re.search(r"Input Voltage\s*\|\s*([\d.]+)\s*V", line)
                current_match = re.search(r"Input Current \s*\|\s*([\d.]+)\s*A", line) 
                if uptime_match:
                    sys_info["Uptime (days)"] = int(uptime_match.group(1))
                    break
                else:
                    sys_info["Uptime (days)"] = 0
                if serial_match:
                     sys_info["Serial Number"] = serial_match.group(1)
                if voltage_match:
                    sys_info[f"Voltage{voltage_index}"] = float(voltage_match.group(1))
                    voltage_index += 1
                if current_match:
                    sys_info[f"Current{current_index}"] = float(current_match.group(1))
                    current_index += 1
        # Extract versions from version file
        with open(version_file, "r") as file:
            content = file.read()
            versions = {
                "UI Version": re.search(r'UI Version:\s*([\d.]+)', content),
                "CLI Version": re.search(r'CLI Version:\s*([\d.]+)', content),
                "SAB Version": re.search(r'SAB Version:\s*([\d.]+)', content)
            }
            
            for name, match in versions.items():
                if match:
                    sys_info[name] = match.group(1)
                else:
                    sys_info[name] = "Not Found"
                    
        # Convert to list of dict format that pandas expects
        return [sys_info]
    except Exception as e:
        print(f"Error extracting system info: {e}")
        return []

# Function to extract host information from SCST configuration
def extract_host_info():
    import os, glob, re, sqlite3
    print("[DEBUG] Starting extract_host_info()")

    scst_dir = "./SCST"
    db_dir = "./Database"

    scst_files = sorted(glob.glob(os.path.join(scst_dir, "scst_2*.conf")), reverse=True)
    if not scst_files:
        print("Error: No 'scst_*.conf' files found in /SCST directory.")
        return []

    input_file = scst_files[0]
    print(f"[DEBUG] SCST file used: {input_file}")

    try:
        with open(input_file, "r") as file:
            lines = file.readlines()

        target_blocks = []
        current_target = None
        brace_count = 0

        # Manually parse TARGET blocks
        for line in lines:
            target_match = re.match(r"\s*TARGET\s+([0-9a-fA-F:]+)\s*\{", line)
            if target_match:
                if current_target:
                    print("[WARNING] Nested TARGET found, skipping previous unfinished block.")
                current_target = {
                    "address": target_match.group(1),
                    "content": [line]
                }
                brace_count = 1
                continue

            if current_target:
                current_target["content"].append(line)
                brace_count += line.count("{")
                brace_count -= line.count("}")
                if brace_count == 0:
                    # Complete block
                    target_blocks.append((current_target["address"], "".join(current_target["content"])))
                    current_target = None

        host_data = []

        for target_address, target_body in target_blocks:
            groups = re.findall(r"GROUP\s+([\w-]+)\s*\{([\s\S]*?)\}", target_body)

            for group_name, group_content in groups:
                luns = re.findall(r"LUN\s+(\d+)\s+([\w_]+)", group_content)
                initiators = re.findall(r"INITIATOR\s+([0-9a-fA-F:]+)", group_content)

                unique_luns = sorted(set(luns), key=lambda x: int(x[0]))
                unique_initiators = sorted(set(initiators))

                sab_db_file = os.path.join(db_dir, "sab.db")
                host_map = {}

                with sqlite3.connect(sab_db_file) as conn:
                    cursor = conn.cursor()

                    for initiator in unique_initiators:
                        cursor.execute("""
                            SELECT initiator_name, to_host_id 
                            FROM hostinitiators
                            WHERE initiator_name = ?
                        """, (initiator,))
                        to_host_row = cursor.fetchone()

                        if to_host_row:
                            to_host_id = to_host_row[1]
                            cursor.execute("""
                                SELECT NAME 
                                FROM host
                                WHERE ID = ?
                            """, (to_host_id,))
                            host_row = cursor.fetchone()
                            host_map[initiator] = host_row[0] if host_row else "Not Found"
                        else:
                            host_map[initiator] = "Not Found"

                # Group initiators and LUNs by host
                host_luns_initiators = {}
                for initiator, host_name in host_map.items():
                    if host_name not in host_luns_initiators:
                        host_luns_initiators[host_name] = {
                            "luns": set(),
                            "initiators": [],
                            "target_map": {}
                        }
                    host_luns_initiators[host_name]["luns"].update(
                        [lun[1] for lun in unique_luns if lun[1] != "device_null"]
                    )
                    host_luns_initiators[host_name]["initiators"].append(initiator)
                    host_luns_initiators[host_name]["target_map"][initiator] = target_address

                for host_name, data in host_luns_initiators.items():
                    for initiator in sorted(data["initiators"]):
                        target = data["target_map"].get(initiator, "")
                        if not data["luns"]:
                            host_data.append({
                                "Access Point": group_name,
                                "Host": host_name,
                                "LUNs": "",
                                "Initiator Addresses": initiator,
                                "Target Address": target
                            })
                        else:
                            for lun in sorted(data["luns"]):
                                host_data.append({
                                    "Access Point": group_name,
                                    "Host": host_name,
                                    "LUNs": lun,
                                    "Initiator Addresses": initiator,
                                    "Target Address": target
                                })

        print(f"[DEBUG] Parsed {len(target_blocks)} target blocks.")
        print(f"[DEBUG] Returning {len(host_data)} host entries.")
        return host_data

    except Exception as e:
        print(f"[ERROR] extract_host_info failed: {e}")
        return []

    except ValueError as e:
        print(f"Error: {e}")
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        if 'conn' in locals():
            conn.close()
def extractor():
    script_dir = os.getcwd()
    passwd = getpass.getpass("Enter ZIP password: ")

    # Find full_log.zip or full_log.log
    log_file = None
    for file in os.scandir(script_dir):
        if 'full_log' in file.name.lower():
            log_file = file.name
            break

    if not log_file:
        print("No log file found.")
        return

    # Rename if needed
    if log_file.lower().endswith('.log'):
        zip_name = log_file[:-4] + '.zip'
        os.rename(log_file, zip_name)
        log_file = zip_name

    # First unzip (password-protected)
    with ZipFile(log_file, 'r') as zip_ref:
        zip_ref.extractall(script_dir, pwd=bytes(passwd, 'utf-8'))

    # Find inner zip
    inner_zip = None
    for file in os.scandir(script_dir):
        if file.name.endswith('.zip') and 'full_log' not in file.name.lower():
            inner_zip = file.name
            break

    if inner_zip:
        with ZipFile(inner_zip, 'r') as zip_ref:
            # Safe extract to avoid absolute paths or path traversal
            for member in zip_ref.namelist():
                member_path = os.path.normpath(member)
                if not member_path.startswith("..") and not os.path.isabs(member_path):
                    target_path = os.path.join(script_dir, member_path)
                    zip_ref.extract(member, script_dir)
            print("Extraction complete.")
    else:
        print("No inner zip file found.")
#Extract files
extractor()
# Get the directory of the script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Path to the smarts.mylinux file in the /SystemOverallInfo directory
smarts_file_path = os.path.join(script_dir, "SystemOverallInfo", "smarts.mylinux")

# Path to the storcli-Sall-show-all.mylinux file in the /SystemOverallInfo directory
storcli_file_path = os.path.join(script_dir, "SystemOverallInfo", "storcli-Sall-show-all.mylinux")

# Read the log files
try:
    with open(smarts_file_path, "r") as file:
        smarts_content = file.read()
    with open(storcli_file_path, "r") as file:
        storcli_content = file.read()
except FileNotFoundError:
    print(f"Error: The required files were not found in the /SystemOverallInfo directory.")
    exit(1)

# Extract SSD, HDD, and device info data
ssd_data = extract_ssd_parameters(smarts_content)
hdd_data = extract_hdd_parameters(smarts_content)
device_data = extract_device_info(smarts_content)

# Extract host information
host_data = extract_host_info()
# Extract General Device Info
sys_info = extract_sysinfo()
# Extract enclosure/slot information
serial_numbers = set([disk["Serial Number"] for disk in ssd_data + hdd_data if disk["Serial Number"] != "Unknown"])
enclosure_slot_data = extract_enclosure_slot_info(storcli_content, serial_numbers)

# Add enclosure/slot information to SSD and HDD data
for disk in ssd_data + hdd_data:
    if disk["Serial Number"] in enclosure_slot_data:
        disk["Enclosure/Slot"] = enclosure_slot_data[disk["Serial Number"]]["enclosure_slot"]
        disk["Shield Counter"] = enclosure_slot_data[disk["Serial Number"]]["shield_counter"]
        disk["Media Error Count"] = enclosure_slot_data[disk["Serial Number"]]["media_error_count"]
        disk["Other Error Count"] = enclosure_slot_data[disk["Serial Number"]]["other_error_count"]
    else:
        disk["Enclosure/Slot"] = ""

# Reorder columns to make "Enclosure/Slot" the first column
def reorder_columns(data):
    return [{"Enclosure/Slot": disk.get("Enclosure/Slot", "N/A"), **disk} for disk in data]

ssd_data = reorder_columns(ssd_data)
hdd_data = reorder_columns(hdd_data)

# Remove rows where "Enclosure/Slot" is "N/A"
ssd_data = [disk for disk in ssd_data]
hdd_data = [disk for disk in hdd_data]

# Create an Excel writer
excel_path = 'smart_data.xlsx'
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Write SMART data to first sheet
    df_smart = pd.DataFrame(ssd_data + hdd_data)

    # Ensure empty rows are preserved in the DataFrame
    # Replace empty strings with NaN (optional, but helps with consistency)
    df_smart.replace("", pd.NA, inplace=True)

    # Write the DataFrame to the Excel file
    df_smart.to_excel(writer, sheet_name="SMART Data", index=False)

    # Write device temperature & powered-up hours to the second sheet (only if non-empty)
    if device_data:
        df_device = pd.DataFrame(device_data)
        df_device.to_excel(writer, sheet_name="Device Info", index=False)

    # Write host information to the third sheet (only if non-empty)
    if host_data:
        df_host = pd.DataFrame(host_data)
        df_host.to_excel(writer, sheet_name="Host Info", index=False)
    if sys_info:
        df_host = pd.DataFrame(sys_info)
        df_host.to_excel(writer, sheet_name="General System Info", index=False)

# Open the Excel file and format it
wb = load_workbook(excel_path)

# Function to merge cells for a specific column
def merge_cells_for_column(ws, col_idx):
    prev_value = None
    start_row = 2  # Start from the second row (first row is headers)

    for row in range(2, ws.max_row + 1):
        current_value = ws.cell(row=row, column=col_idx).value
        if current_value == prev_value:
            continue
        if prev_value is not None and start_row is not None:
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=row-1, end_column=col_idx)
            ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")
        prev_value = current_value
        start_row = row

    if prev_value is not None and start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")

# Function to adjust column widths automatically
def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

# Format all sheets except "Device Info"
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if sheet_name != "Device Info":  # Skip merging for "Device Info" sheet
        merge_cells_for_column(ws, 1)  # Merge "Enclosure/Slot" column (column 1)
        merge_cells_for_column(ws, 2)  # Merge "Serial Number" column (column 2)
        merge_cells_for_column(ws, 3)  # Merge "Device Model" column (column 3)
    adjust_column_widths(ws)  # Adjust column widths for all sheets
if "Host Info" in wb.sheetnames: 
    host_info_sheet = wb["Host Info"]
    merge_cells_for_column(host_info_sheet, 4)  # Merge "Initiators" column (column 4)
    merge_cells_for_column(host_info_sheet, 5)  # Merge "Targets" column (column 5) 
wb.save(excel_path)
print("SMART data, device info, and host info extracted and written to smart_data.xlsx with proper formatting.")
