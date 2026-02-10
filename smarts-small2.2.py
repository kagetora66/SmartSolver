from contextlib import aclosing
from datetime import datetime
import json
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
import glob
import sqlite3
import subprocess
from openpyxl.styles import PatternFill
from pathlib import Path
import time
from collections import defaultdict
from collections import Counter
from openpyxl.styles import Border, Side
# Define required parameters
def merge_cells_for_column(ws, col_idx):
    prev_value = None
    start_row = 2  # Start from the second row (first row is headers)
    for row in range(2, ws.max_row + 1):
        current_value = ws.cell(row=row, column=col_idx).value
        if current_value == prev_value:
            continue
        if prev_value is not None and start_row is not None:
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=row-1, end_column=col_idx)
            ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")
        prev_value = current_value
        start_row = row

    if prev_value is not None and start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")
def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = get_column_letter(col[0].column)
        if max_length < 70:
            ws.column_dimensions[col_letter].width = max_length + 4
        else:
            ws.column_dimensions[col_letter].width = 70
            for column in ws.iter_cols():
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, vertical='center') 
def adjust_height(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            ws.row_dimensions[cell.row].height = 80

def log_extract(log_path):
  with open(log_path, "rb") as file:
        content = file.read()
        content_clean = content.replace(b'\x00', b'')
        content_uft8 = content_clean.decode('ascii', errors='ignore')
        content_excel_safe = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]', '', content_uft8)
        lines = content_excel_safe.splitlines() 
        recent_lines = lines[-1000:]
        if "dmesg" in log_path:
            return {"Logs": recent_lines}
        # Process lines to group multi-line entries
        processed_logs = []
        current_entry = []
        
        for line in recent_lines:
            line = line.rstrip()  # Remove trailing whitespace/newline
            # Check if line starts with a timestamp (format: YYYY-MM-DD HH:MM:SS,mmm)
            if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3}', line):
                # If we have a current entry, save it before starting new one
                if current_entry:
                    processed_logs.append(" ".join(current_entry))
                    current_entry = []
                current_entry.append(line)
            else:
                # Continuation line - add to current entry
                if current_entry:
                    current_entry.append(line)
        
        # Don't forget the last entry
        if current_entry:
            processed_logs.append(" ".join(current_entry))
        
        return {"Log": processed_logs}

ssd_params = [
    "Reallocated_Sector_Ct",
    "Power_On_Hours",
    "Wear_Leveling_Count",
    "Used_Rsvd_Blk_Cnt_Tot",
    "Runtime_Bad_Block",
    "Reported_Uncorrect",
    "Hardware_ECC_Recovered",
    "Total_LBAs_Written",
    "Total Size Written (TB)"
]
hdd_params = [
    "Elements in grown defect list",
    "Total Uncorrected Errors",
    "Accumulated start-stop cycles",
    "Accumulated load-unload cycles"
]
micron_ssd_params = {
    "Raw_Read_Error_Rate": "1",
    "Reallocated_Sector_Ct": "5",
    "Reported_Uncorrect": "187",
    "Hardware_ECC_Recovered": "195",
    "Unused_Rsvd_Blk_Cnt_Tot": "180",
    "Total_LBAs_Written": "246"  # Unknown attribute at ID# 246
}

threshold_sata_ssd = {
    "Reallocated_Sector_Ct": "10",
    "Power_On_Hours": "5",
    "Wear_Leveling_Count": "10",
    "Used_Rsvd_Blk_Cnt_Tot": "10",
    "Runtime_Bad_Block": "10",
    "Reported_Uncorrect": "97",
    "Hardware_ECC_Recovered": "100",
    "Total Size Written (TB)": "6400"
} 

threshold_micron_ssd = {
    "Raw_Read_Error_Rate": "50",
    "Reallocated_Sector_Ct": "1",
    "Reported_Uncorrect": "0",
    "Hardware_ECC_Recovered": "0",
    "Unused_Rsvd_Blk_Cnt_Tot": "0",
    "Total Size Written (TB)": "6400"
    
}

smart_pattern = re.compile(
    r"(\d+)\s+([\w_]+)\s+0x[0-9a-fA-F]+\s+(\d+)\s+\d+\s+(\d+)\s+\w+-?\w*\s+\w+\s+-\s+(\d+)"
)
def extract_ssd_parameters(log_content):
    data = []
    # Split the log into blocks for each disk.
    disk_blocks = re.findall(
        r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)",
        log_content, re.DOTALL
    )
    
    for block in disk_blocks:
        # Determine disk type based on certain strings
        is_ssd = not re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        is_sas_ssd = re.search(r"Transport protocol:\s+SAS", block, re.IGNORECASE)
        is_micron_ssd = re.search(r"Device Model:\s+Micron", block, re.IGNORECASE)
        disk_type = ""
        tb_value = ""
        capacity_match = re.search(r"\[([^\]]+)\]", block)
        if capacity_match:
            tb_value = capacity_match.group(1)
            user_capacity = float(re.search(r'\d+(?:\.\d+)?', tb_value).group()) if re.search(r'\d+(?:\.\d+)?', tb_value) else 0            
        if is_ssd:
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            model_match =  re.search(r"Device Model:\s+(.+)", block, re.IGNORECASE)
            device_model = model_match.group(1) if model_match else "Unknown" 
            if is_sas_ssd:
                # --- SAS SSD extraction (existing logic) ---
                elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
                start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
                load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)
                read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
                write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
                endurance_indicator = re.search(r"SS Media used endurance indicator:\s+(\d+)", block)
                model_match_sas = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)
                device_model = model_match_sas.group(1) if model_match_sas else "Unknown" 
                brand = "Samsung"
                disk_type = "SSD SAS"
                read_error_value = int(read_error_match.group(1)) if read_error_match else 0
                write_error_value = int(write_error_match.group(1)) if write_error_match else 0
                total_uncorrected_errors = read_error_value + write_error_value
                total_uncorrrected_threshold = int(user_capacity)  * 5
                Element_in_defect_threshold = int(user_capacity)  * 100
        #We convert to str again for consistency
                threshold_sas_ssd = {
                    "Elements in grown defect list": str(Element_in_defect_threshold),
                    "Total Uncorrected Errors": str(total_uncorrrected_threshold),
                    "Accumulated start-stop cycles": "10000",
                    "Accumulated load-unload cycles": "300000",
                    "SS Media used endurance indicator %": "90"
                    }
                hdd_values = [
                    ("Elements in grown defect list", elements_grown_defect),
                    ("Total Uncorrected Errors", total_uncorrected_errors),
                    ("Accumulated start-stop cycles", start_stop_cycles),
                    ("SS Media used endurance indicator %", endurance_indicator)
                ]

                for param, match in hdd_values:
                    if isinstance(match, int):  # For Total Uncorrected Errors
                        raw_value = str(match)
                    elif match:
                        raw_value = match.group(1)
                    else:
                        continue
                    threshold = threshold_sas_ssd.get(param, "-")
                    data.append({
                        "Brand" : brand,
                        "Device Model": device_model, 
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": param,
                        "Threshold" : threshold,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
                    
            else:
                # --- Non-SAS SSD extraction ---
                # smart_pattern should be defined elsewhere to extract tuples:
                # (attr_id, attr_name, value, raw_value)
                smart_matches = smart_pattern.findall(block)
                total_lba_written = None  # To store Total_LBAs_Written when found
                disk_type = "SSD SATA"
                if is_micron_ssd:
                    brand = "Micron"
                    # Define the list of expected Micron SMART parameters.
                    micron_params = [
                        "Raw_Read_Error_Rate", 
                        "Reallocated_Sector_Ct", 
                        "Reported_Uncorrect", 
                        "Hardware_ECC_Recovered", 
                        "Unused_Rsvd_Blk_Cnt_Tot",
                        "Used_Reserve_Block_Count"
                    ]

                    for match in smart_matches:
                        attr_id, attr_name, value, threshold, raw_value = match
                        # Check for the known Micron parameters.
                        if attr_name in micron_params:
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": attr_name,
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        # Detect Total_LBAs_Written by its attribute ID "246"
                        elif attr_id.strip() == "246":
                            try:
                                total_lba_written = int(raw_value)
                            except ValueError:
                                total_lba_written = None
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": "Total_LBAs_Written",
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        #Detect used block count
                        elif attr_id.strip() == "170":
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": "Used_Reserve_Block_Count",
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        elif attr_id.strip() == "202":
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": "Percent Life Time Remaining",
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        elif attr_id.strip() == "9":
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": "Power_On_Hours",
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                    if total_lba_written is not None:

                        total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                        threshold_write_micron = 365 * 0.8 * 5 * user_capacity
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)",
                            "Threshold": str(round(threshold_write_micron, -3)),
                            "Value": "-",
                            "Raw Value": f"{total_size_written_tb:.2f}"
                        })

                        total_size_micron_waf2 = total_size_written_tb * 2
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-WAF(2)",
                            "Threshold": str(round(threshold_write_micron, -1)),
                            "Value": "-",
                            "Raw Value": f"{total_size_micron_waf2:.2f}"
                        })
                        total_size_micron_waf4 = total_size_written_tb * 4
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-WAF(4)",
                            "Threshold": str(round(threshold_write_micron, -1)),
                            "Value": "-",
                            "Raw Value": f"{total_size_micron_waf4:.2f}"
                        })

                else:
                    brand = "SAMSUNG"
                    disk_type = "SSD SATA"
                    # Existing logic for non-Micron SATA SSDs
                    for match in smart_matches:
                        attr_id, attr_name, value, threshold, raw_value = match
                        if attr_name in ssd_params: # ssd_params defined elsewhere
                            if attr_name == "Total_LBAs_Written":
                                try:
                                    total_lba_written = int(raw_value)
                                except ValueError:
                                    total_lba_written = None
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": attr_name,
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        if attr_name == "Wear_Leveling_Count":
                            wearlevel = raw_value
                    if total_lba_written is not None:
                        total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                        if "MZ" in device_model:
                            threshold_samsung_write = round(365 * 5 * user_capacity * 4, -3)
                        elif "850" in device_model:
                            threshold_samsung_write = 150
                            user_capacity = 0.25
                        elif "860" in device_model:
                            threshold_samsung_write = 300
                            user_capacity = 0.25
                        elif "870" in device_model:
                            threshold_samsung_write = 150
                            user_capacity = 0.25
                        else:
                            threshold_samsung_write = 300 #undefined
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)",
                            "Threshold": str(threshold_samsung_write),
                            "Value": "-",
                            "Raw Value": f"{total_size_written_tb:.2f}"
                        })
                        total_size_written_ssd = float(wearlevel) * user_capacity
                        total_size_samsung_waf2 = total_size_written_tb * 2
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-WAF(2)",
                            "Threshold": str(threshold_samsung_write),
                            "Value": "-",
                            "Raw Value": f"{total_size_samsung_waf2:.2f}"
                        })
                        total_size_samsung_waf4 = total_size_written_tb * 4
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-WAF(4)",
                            "Threshold": str(threshold_samsung_write),
                            "Value": "-",
                            "Raw Value": f"{total_size_samsung_waf4:.2f}"
                        })
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)-SSD_Ctl",
                            "Threshold": str(threshold_samsung_write),
                            "Value": "-",
                            "Raw Value": f"{total_size_written_ssd:.2f}"
                        })
            # Add an empty row after each disk's data for readability.
            data.append({
                "Brand": "",
                "Device Model": "",
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
            for i in range(len(data) - 2):
                if (
                    data[i]["Parameter"] == "Total_LBAs_Written" and
                    data[i + 1]["Parameter"] == "Unused_Rsvd_Blk_Cnt_Tot"
                ):
                    # Swap their positions
                    data[i], data[i + 1] = data[i + 1], data[i]
            
    return data

# Function to extract HDD parameters
def extract_hdd_parameters(log_content):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    for block in disk_blocks:
        capacity_match = re.search(r"\[([^\]]+)\]", block)
        if capacity_match:
            tb_value = capacity_match.group(1)
        user_capacity = float(re.search(r'\d+(?:\.\d+)?', tb_value).group()) if re.search(r'\d+(?:\.\d+)?', tb_value) else 0
        total_uncorrrected_threshold = int(user_capacity)  * 5
        Element_in_defect_threshold = int(user_capacity)  * 100
        #We convert to str again for consistency
        threshold_hdd_sata = {
                "Elements in grown defect list": str(Element_in_defect_threshold),
                "Total Uncorrected Errors": str(total_uncorrrected_threshold),
                "Accumulated start-stop cycles": "10000",
                "Accumulated load-unload cycles": "300000"
                }
        is_sata = re.search(r'\bSATA\b', block, re.IGNORECASE)         
        is_hdd = re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        if is_hdd and not is_sata:

            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            disk_type = "HDD SAS"
            model_match =  re.search(r"Device Model:\s+(.*?)\s*$", block, re.IGNORECASE)
            model_match_hp = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)
            device_model = (
                    model_match.group(1) if model_match 
                    else model_match_hp.group(1) if model_match_hp
                    else "Unknown"
                    )
            #seagate_match =  re.search(r"Vendor:\s+(\S+)", block, re.IGNORECASE)
            if device_model.startswith("ST"):
                brand = "SEAGATE"
            elif device_model.startswith("TOSHIBA"):
                brand = "Toshiba"
            else:
                brand = "HP"

  
            elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
            start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
            load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)

            # Extract Total Uncorrected Errors
            read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
            write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
            
            read_error_value = int(read_error_match.group(1)) if read_error_match else 0
            write_error_value = int(write_error_match.group(1)) if write_error_match else 0
            total_uncorrected_errors = read_error_value + write_error_value

            # Sort the parameters in the desired order
            hdd_values = [
                ("Elements in grown defect list", elements_grown_defect),
                ("Total Uncorrected Errors", total_uncorrected_errors),
                ("Accumulated start-stop cycles", start_stop_cycles),
                ("Accumulated load-unload cycles", load_unload_cycles)
            ]
            
            for param, match in hdd_values:
                if isinstance(match, int):  # For Total Uncorrected Errors
                    raw_value = str(match)
                    threshold = threshold_hdd_sata.get(param, "-")
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": param,
                        "Threshold": threshold,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
                elif match:
                    raw_value = match.group(1) if hasattr(match, "group") else str(match)
                    threshold = threshold_hdd_sata.get(param, "-")
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": param,
                        "Threshold": threshold,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
            # Add an empty row after each disk's data
            data.append({
                "Brand": "",
                "Device Model": "",
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
        elif is_hdd and is_sata:
            # --- SATA HDD extraction ---
            # smart_pattern should be defined elsewhere to extract tuples:
            # (attr_id, attr_name, value, raw_value)
            smart_matches = smart_pattern.findall(block) 
            total_lba_written = None  # To store Total_LBAs_Written when found    
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            disk_type = "HDD SATA"
            model_match = re.search(r"Device Model:\s+(.*?)\s*$", block, re.IGNORECASE | re.MULTILINE)
            model_match_hp = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)

            #seagate_match =  re.search(r"Vendor:\s+(\S+)", block, re.IGNORECASE)
            #brand = seagate_match.group(1) if seagate_match else "HP"
            device_model = (
                    model_match.group(1) if model_match 
                    else model_match_hp.group(1) if model_match_hp
                    else "Unknown"
                    )
            if device_model.startswith("ST"):
                brand = "SEAGATE"
            elif device_model.startswith("TOSHIBA"):
                brand = "Toshiba"
            # Define the list of expected Micron SMART parameters.
            hdd_sata_params = [
                "Raw_Read_Error_Rate", 
                "Spin_Retry_Count", 
                "Reported_Uncorrect",
                "Command_Timeout", 
                "Current_Pending_Sector",
                "Offline_Uncorrectable",
                "Multi_Zone_Error_Rate"
            ]
            for match in smart_matches:
                attr_id, attr_name, value, threshold, raw_value = match
                # Check for the known SATA parameters.
                if attr_name in hdd_sata_params:
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": attr_name,
                        "Threshold": threshold,
                        "Value": value,
                        "Raw Value": raw_value
                    })
                    # Detect Total_LBAs_Written by its attribute ID "241"
                elif attr_id.strip() == "241":
                    try:
                        total_lba_written = int(raw_value)
                    except ValueError:
                        total_lba_written = None
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": "Total_LBAs_Written",
                        "Threshold": threshold,
                        "Value": value,
                        "Raw Value": raw_value
                    })
                    
            if total_lba_written is not None:
                total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                threshold = threshold_micron_ssd.get("Total Size Written (TB)", "-")
                data.append({
                    "Brand": brand,
                    "Device Model": device_model,
                    "Serial Number": serial_number,
                    "Interface": disk_type,
                    "Size": tb_value,
                    "Parameter": "Total Size Written (TB)",
                    "Threshold": threshold,
                    "Value": "-",
                    "Raw Value": f"{total_size_written_tb:.2f}"
                    })

            data.append({
                "Brand": "",
                "Device Model": "",
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
    return data
#Needs work
def parse_scst_to_dict(log_content):
    """Parse SCST config file and return list of dictionaries.""" 
    with open(log_content, 'r') as f:
        lines = f.readlines()
    target_blocks = []
    current_target = None
    brace_count = 0

        # Manually parse TARGET blocks
    for line in lines:
        target_match = re.match(r"\s*TARGET\s+(.+)\s*\{", line)
        if target_match:
            if current_target:
                print("[WARNING] Nested TARGET found, skipping previous unfinished block.")
            current_target = {
                "address": target_match.group(1),
                "content": [line]
            }
            brace_count = 1
            continue

            if current_target:
                current_target["content"].append(line)
                brace_count += line.count("{")
                brace_count -= line.count("}")
                if brace_count == 0:
                    # Complete block
                    target_blocks.append((current_target["address"], "".join(current_target["content"])))
                    current_target = None
        for target_address, target_body in target_blocks:
            groups = re.findall(r"GROUP\s+([\w-]+)\s*\{([\s\S]*?)\}", target_body)

            for group_name, group_content in groups:
                # Handle empty Access Control
                access_control = group_name if group_name else "-"
                
                luns = re.findall(r"LUN\s+(\d+)\s+([\w_]+)", group_content)
                initiators = re.findall(r"INITIATOR\s+(.+)", group_content)

                unique_luns = sorted(set(luns), key=lambda x: int(x[0]))
                unique_initiators = sorted(set(initiators))

def extract_device_info(log_content, enclosure):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    tmp_data = []
    for block in disk_blocks:
         serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
         serial_number = serial_match.group(1) if serial_match else None

         temp_match = re.search(r"Current Drive Temperature:\s+(\d+)", block)
         hours_match = re.search(r"number of hours powered up\s+=\s+([\d.]+)", block)
         #For Samsung SSDs
         temp_match_sam = re.search(r"194\s+[\w_]+\s+[\w\d]+\s+\d+\s+\d+\s+\d+\s+\w+\s+\w+\s+-\s+(\d+)", block)
         if not temp_match_sam: temp_match_sam = re.search(r"190\s+[\w_]+\s+[\w\d]+\s+\d+\s+\d+\s+\d+\s+\w+\s+\w+\s+-\s+(\d+)", block) 
         hours_match_sam = re.search(r"9\s+[\w_]+\s+[\w\d]+\s+\d+\s+\d+\s+\d+\s+\w+\s+\w+\s+-\s+(\d+)", block)
         tsw = "-"
         tsw_ctl = "-"
         hours = "-"
         tsw_waf2 = "-"
         tsw_waf4 = "-"
         percent_life = "-"
         slot = "-"
         model = "-"
         state = "-"
         ss_media = "-"
         if serial_number:
             if temp_match:
                 temperature = f"{temp_match.group(1)}"
             else:
                 temperature = "-"

             for enc in enclosure:
                 if not enc.get("En/Slot"):
                     enc["En/Slot"] = ""
                 if not enc.get("Disk State"):
                     enc["Disk State"] = ""
                 if serial_number in enc["Serial Number"]:
                     interface = enc["Interface"]
                     slot = enc["En/Slot"]
                     model = enc["Device Model"]
                     state = enc["Disk State"]
                     if enc["Parameter"] == "Total Size Written (TB)":
                         tsw = enc["Raw Value"]
                         tsw_waf2 = float(tsw) * 2
                         tsw_waf4 = float(tsw) * 4
                     elif enc["Parameter"] == "Total Size Written (TB)-SSD_Ctl":
                         tsw_ctl = enc["Raw Value"]
                     if enc["Parameter"] == "Power_On_Hours":
                         hours = enc["Raw Value"]
                     if enc["Parameter"] == "Percent Life Time Remaining":
                         percent_life = enc["Raw Value"]
                     if enc["Parameter"] == "SS Media used endurance indicator %":
                         ss_media = enc["Raw Value"]
             if hours_match:
                 hours = hours_match.group(1)
             tmp_data.append({
                "Enc/Slot": slot,
                "Interface": interface,
                "Model": model,
                "Serial Number": serial_number,
                "Temp": temperature,
                "Powered Up Hours": hours,
                "TBW": tsw,
                "TBW-WAF2": tsw_waf2,
                "TBW-WAF4": tsw_waf4,
                "TBW (Ctl)": tsw_ctl,
                "Percent Life": percent_life,
                "Endurance": ss_media
                })

         elif serial_number:
             if temp_match_sam:
                temperature = f"{temp_match_sam.group(1)}"
             else:
                temperature = "-"
             if hours_match_sam:
                 hours = hours_match_sam.group(1)
             data.append({
                 "Enc/Slot": slot,
                 "Interface": interface,
                 "Model": model,
                 "Serial Number": serial_number,
                 "Temp": temperature,
                 "Powered Up Hours": hours,
                 "TBW": tsw,
                 "TBW-WAF2": tsw_waf2,
                 "TBW-WAF4": tsw_waf4,
                 "TBW (Ctl)": tsw_ctl,
                 "Percent Life": percent_life,
                 "Endurance": ss_media
                 })

    data.extend(tmp_data)
    return data

def extract_sysinfo():
    
    pmc_file = glob.glob(os.path.join(script_dir, "output.txt"))
    sys_info = {}
    if pmc_file:
         # Extract versions from pmc output
         print("System Info extracted from output.txt")
         with open(pmc_file[0], "r") as file:
             content = file.read()
             basic_info = {
                 "SAB ID": re.search(r'hostname\s*:\s*(.*)$', content, re.IGNORECASE | re.MULTILINE), 
                 "SAB Ver": re.search(r'#SAB version\s+([^\s]+)', content),
                 "Rep Ver": re.search(r'REPLICATION VERSION:\s*VERSION=([^\s]+)', content, re.IGNORECASE),
                 "Rapid Ver": re.search(r'Rapidtier Version:\s*([^\s]+)', content),
                 "UI Ver": re.search(r'version\s*=\s*"([^"]+)"', content),
                 "CLI Ver": re.search(r'CLI Version:\n(.+)', content),
                 "BBU": re.search(r'BBU Status =\s*([^\s]+)', content),
                 "CC Status": re.search(r'CC is\s*(.+)', content),
                 "RC Model": re.search(r'Model\s+=\s+(.+)', content),
                 "FW Ver": re.search(r'Firmware Version\s+=\s(.+)', content)
             }
    for name, match in basic_info.items():
        if match:
            sys_info[name] = match.group(1)
        else:
            sys_info[name] = "Not Found"
    return [sys_info]

 # Function to extract enclosure/slot information
def extract_enclosure_slot_info(log_content, serial_numbers):
    enclosure_slot_data = {}

    # Split the log content into lines
    lines = log_content.splitlines()

    current_shield_counter = None
    current_media_error_count = None
    current_other_error_count = None
    predictive_failure_count = None
    disk_status = None
    for i, line in enumerate(lines):
        line = line.strip()

        # Capture Shield Counter
        if line.startswith("Shield Counter"):
            current_shield_counter = line.split(":")[1].strip()
            if current_shield_counter == "0":
                current_shield_counter = "-"

        # Capture Media Error Count
        elif line.startswith("Media Error Count"):
            current_media_error_count = line.split(":")[1].strip()
            if current_media_error_count == "0":
                current_media_error_count = "-"


        # Capture Other Error Count
        elif line.startswith("Other Error Count"):
            current_other_error_count = line.split(":")[1].strip()
            if current_other_error_count == "0":
                current_other_error_count = "-"

        #Capture Predictive Failure Count
        elif line.startswith("Predictive Failure Count"):
            predictive_failure_count = line.split(":")[1].strip()
            if predictive_failure_count == "0":
               predictive_failure_count = "-"

        #Capture disk state 
        
        if " UGood " in line:
            disk_status = "No Config"
        elif "Online" in line:
            disk_status = "Operational"
        elif "Hotspare" in line:
            disk_status = "Hotspare"
        elif " UBad " in line:
            disk_status = "BAD UNCONFIGURED"

        serial_pattern = r'Inquiry Data:\s+([A-Z-0-9]+)\s+[A-Z-0-9]+\s+([A-Z-0-9]+)'
        serial_match = re.search(serial_pattern, line)
        serial = "-"
        #Test
        if line.startswith("Inquiry Data"):
            serial_parts = line.split(" ")
            if "SEAGATE" in serial_parts:
                serial = serial_parts[-1].split("E003")[1]
            elif "samsung" in list(map(str.lower, serial_parts)):
                serial = serial_parts[2]
            else:
                #Micron
                serial = serial_parts[10].split("Micron")[0]
        #if serial_match:
        #    if serial_match.group(1) == "SEAGATE":
        #        serial = serial_match.group(2)
        #        serial = serial.split("E003")[1]
        #    else:
        #        serial = serial_match.group(1)
        # Detect Serial Number
        for srl in serial_numbers:
            if serial in srl:
                # Search backward for Drive line
                for j in range(i, max(i - 29, -1), -1):
                    drive_line = lines[j].strip()
                    enc_pattern = r'Enclosure Device ID:\s+([0-9]+)'
                    slot_pattern = r'Slot Number:\s+([0-9]+)'
                    enc_match = re.search(enc_pattern, drive_line)
                    slot_match = re.search(slot_pattern, drive_line)
                    if enc_match:
                        enclosure = enc_match.group(1)
                    if slot_match:
                        slot = slot_match.group(1)
                        # Now assign all gathered info
                enclosure_slot_data[serial] = {
                    "enclosure_slot": f"{enclosure}/{slot}",
                    "shield_counter": current_shield_counter,
                    "media_error_count": current_media_error_count,
                    "other_error_count": current_other_error_count,
                    "predictive_failure_count": predictive_failure_count,
                    "Disk State": disk_status
                }
    return enclosure_slot_data

def about_info():
    about = []
    script_ver = os.path.basename(__file__)
    about.append({
        "Script Version" : script_ver
        })
    return about

def reorder_columns(data):
    return [{"En/Slot": disk.get("En/Slot", "N/A"), **disk} for disk in data]

#Extract slot number
def extract_slot_port_info():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pmc_files = glob.glob(os.path.join(script_dir, "output.txt"))

    if pmc_files:
        print("output.txt used in extracting slot info")
        with open(pmc_files[0], "r") as f:
            log = f.read()

        slot_data = defaultdict(lambda: {
            'enabled_ports': [],
            'total_ports': 0,
            'port_type': '',
            'ports': {}
        })
    
        current_slot = None
        current_type = None
        current_port = None

        for line in log.splitlines():
            line = line.strip()

            # Detect slot type
            if "--------ISCSI CARDS SLOTS" in line:
                current_type = 'iscsi'
            elif "--------FC HBA CARDS SLOTS" in line:
                current_type = 'fc'

            # Match slot/port headers
            match = re.search(r'(NIC|FC) CARD in SLOT\s+(\d+)\s+PORT\s+(\d+)(?:\s+\([^)]+\))?', line)
            if match:
                current_slot = match.group(2)
                current_port = int(match.group(3))
                slot_info = slot_data[current_slot]
                slot_info['port_type'] = current_type
                slot_info['total_ports'] += 1
                slot_info['ports'][current_port] = {
                    'wwn': '',
                    'connection_type': '-',
                    'speed': '-'
                }
                continue

            # FC ports
            if current_type == 'fc':
                if line.startswith("port_speed"):
                    speed = line.split("=", 1)[1].strip()
                    slot_data[current_slot]['ports'][current_port]['speed'] = speed if "Unknown" not in speed else "-"
                elif line.startswith("wwn ="):
                    wwn = line.split("=", 1)[1].strip().lower().replace("0x", "")
                    wwn_formatted = ":".join(wwn[i:i+2] for i in range(0, len(wwn), 2))
                    slot_data[current_slot]['ports'][current_port]['wwn'] = wwn_formatted
                elif line.startswith("port_type"):
                    port_type = line.split("=", 1)[1].strip()
                    if "NPort" in port_type:
                        slot_data[current_slot]['ports'][current_port]['connection_type'] = "SAN-Switch"
                    elif "LPort" in port_type:
                        slot_data[current_slot]['ports'][current_port]['connection_type'] = "LPort (private loop)"
                    elif "Point" in port_type:
                        slot_data[current_slot]['ports'][current_port]['connection_type'] = "Point-to-Point"

            # iSCSI ports
            elif current_type == 'iscsi':
                if line.startswith("speed_interface"):
                    speed = line.split("=", 1)[1].strip()
                    slot_data[current_slot]['ports'][current_port]['speed'] = speed if "Unknown" not in speed else "-"
                elif line.startswith("mac_address"):
                    mac = line.split("=", 1)[1].strip()
                    slot_data[current_slot]['ports'][current_port]['mac_address'] = mac
                elif line.startswith("iqn."):
                    iqn = line.strip()
                    slot_data[current_slot]['ports'][current_port]['wwn'] = iqn
        result = []
        for slot, info in slot_data.items():
            result.append({
                'slot': f"Slot{slot}",
                'port_type': info['port_type'],
                'total_ports': info['total_ports'],
                'ports': info['ports']
            })

        return result
# Function to extract host information from SCST configuration
def extract_host_info():
    scst_dir = "./"
    db_dir = "./"
    # Checks for new scst file inside script directory
    new_scst_matches = glob.glob(os.path.join(script_dir, "scst.*"))
    if new_scst_matches:
        scst_files = new_scst_matches[0]
        input_file = scst_files
    else:
        scst_files = sorted(glob.glob(os.path.join(scst_dir, "scst.conf")), reverse=True)
        input_file = scst_files[0] if scst_files else None
    
    if not input_file:
        print("Error: No 'scst_*.conf' files found in directory.")
        return []

    # Checks for output.txt file inside script directory 
    target_port_type = {}
    is_pmc = os.path.isfile(os.path.join(os.path.dirname(__file__), 'output.txt'))
    if is_pmc:
        pmc_output = "output.txt"
        current_wwn = None
        print("output.txt used in extracting host info")
        with open(pmc_output, "r") as file:
            lines = file.readlines()
        # We map port connection to wwn addresses 
        for line in lines:
            line = line.strip()
            if line.lower().startswith("wwn = 0x"):
                hex_str = line.split('=')[1].strip().lower().replace('0x', '')
                current_wwn = ':'.join(hex_str[i:i+2] for i in range(0, len(hex_str), 2))
            elif 'Point' in line and current_wwn:
                port_type = "Point to Point"
                target_port_type[current_wwn] = port_type
                current_wwn = None
            elif 'NPort' in line and current_wwn:
                port_type = "SAN Switch"
                target_port_type[current_wwn] = port_type
                current_wwn = None
            elif 'LPort' in line and current_wwn:
                port_type = "LPort (private loop)"
                target_port_type[current_wwn] = port_type
                current_wwn = None
                
    else:
        systemstat_dir = os.path.join(script_dir, "Logs")
        systemstat_file = sorted(glob.glob(os.path.join(systemstat_dir, "system_status_20*.txt")), reverse=True)
        input_sys = systemstat_file[0] if systemstat_file else None
        if input_sys:
            with open(input_sys, 'r') as f:
                content = f.read()
                if content.strip(): 
                    data = json.loads(content)
                    fc_data = data.get('SAB', {}).get('fc_cards', {})
                    for card in fc_data:
                        fc_ports = card.get('fc_port', [])
                        for port in fc_ports:
                            wwn = port.get('wwn')
                            port_type = port.get('type')
                            if 'NPort' in port_type:
                                port_type = "SAN Switch(system_status)"
                            elif 'Point' in port_type:
                                port_type = "Point to Point(system_status)"
                            if wwn and port_type:
                                target_port_type[wwn] = port_type.strip()  # .strip() to remove newlines
    try:
        with open(input_file, "r") as file:
            lines = file.readlines()

        target_blocks = []
        current_target = None
        brace_count = 0

        # Manually parse TARGET blocks
        for line in lines:
            target_match = re.match(r"\s*TARGET\s+(.+)\s*\{", line)
            if target_match:
                if current_target:
                    print("[WARNING] Nested TARGET found, skipping previous unfinished block.")
                current_target = {
                    "address": target_match.group(1),
                    "content": [line]
                }
                brace_count = 1
                continue

            if current_target:
                current_target["content"].append(line)
                brace_count += line.count("{")
                brace_count -= line.count("}")
                if brace_count == 0:
                    # Complete block
                    target_blocks.append((current_target["address"], "".join(current_target["content"])))
                    current_target = None

        host_data = []

        for target_address, target_body in target_blocks:
            groups = re.findall(r"GROUP\s+([\w-]+)\s*\{([\s\S]*?)\}", target_body)

            for group_name, group_content in groups:
                # Handle empty Access Control
                access_control = group_name if group_name else "-"
                
                luns = re.findall(r"LUN\s+(\d+)\s+([\w_]+)", group_content)
                initiators = re.findall(r"INITIATOR\s+(.+)", group_content)

                unique_luns = sorted(set(luns), key=lambda x: int(x[0]))
                unique_initiators = sorted(set(initiators))

                sab_db_file = os.path.join(db_dir, "sab.db")
                host_map = {}
                lun_name_map = {}  # Dictionary to map backend LUN names to frontend names
                if os.path.exists(sab_db_file):
                    with sqlite3.connect(sab_db_file) as conn:
                        cursor = conn.cursor()
                        for initiator in unique_initiators:
                            cursor.execute("""
                                SELECT initiator_name, to_host_id 
                                FROM hostinitiators
                                WHERE initiator_name = ?
                            """, (initiator,))
                            to_host_row = cursor.fetchone()

                            if to_host_row:
                                to_host_id = to_host_row[1]
                                cursor.execute("""
                                    SELECT NAME 
                                    FROM host
                                    WHERE ID = ?
                                """, (to_host_id,))
                                host_row = cursor.fetchone()
                                host_map[initiator] = host_row[0] if host_row else "Not Found"
                            else:
                                host_map[initiator] = "Not Found"
                        for lunid, lun_be_name in unique_luns:
                            cursor.execute(
                                'SELECT "fe_name" FROM lun_name_mapper WHERE "be_name" = ?',
                                (lun_be_name,)
                            )
                            lun_fe_row = cursor.fetchone()
                            if lun_fe_row:
                                lun_name_map[lun_be_name] = lun_fe_row[0]
                            else:
                                lun_name_map[lun_be_name] = lun_be_name
                else:
                    print("SAB DB not found")
                    return []

                # Group initiators and LUNs by host
                host_luns_initiators = {}
                for initiator, host_name in host_map.items():
                    if host_name not in host_luns_initiators:
                        host_luns_initiators[host_name] = {
                            "luns": set(),
                            "initiators": [],
                            "target_map": {}
                        }
                    host_luns_initiators[host_name]["luns"].update(
                        [lun[1] for lun in unique_luns if lun[1] != "device_null"]
                    )
            # Add LUNs with frontend names if available
                    for lun_id, lun_be_name in unique_luns:
                        if lun_be_name != "device_null":
                            lun_fe_name = lun_name_map.get(lun_be_name, lun_be_name)
                            host_luns_initiators[host_name]["luns"].add(lun_fe_name)
                            if lun_be_name != lun_fe_name:
                                host_luns_initiators[host_name]["luns"].remove(lun_be_name)

                    host_luns_initiators[host_name]["initiators"].append(initiator)
                    host_luns_initiators[host_name]["target_map"][initiator] = target_address
                for host_name, data in host_luns_initiators.items():
                    host_name = host_name if host_name else "-"
                    for initiator in sorted(data["initiators"]):
                        initiator = initiator if initiator else "-"
                        target = data["target_map"].get(initiator, "-")
                        port_type = target_port_type.get(target.strip(), "na") if target_port_type else "-"
                        
                        if not data["luns"]:
                            host_data.append({
                                "Access Control": access_control,
                                "Host": host_name,
                                "LUNs": "-",
                                "Initiator Addresses": initiator,
                                "Target Address": target,
                                "Connection Type": port_type
                            })
                        else:
                            for lun in sorted(data["luns"]):
                                host_data.append({
                                    "Access Control": access_control,
                                    "Host": host_name,
                                    "LUNs": lun,
                                    "Initiator Addresses": initiator,
                                    "Target Address": target,
                                    "Connection Type": port_type
                                })

        if not host_data:
            if target_port_type:
                for target, port_type in target_port_type.items():
                    host_data.append({
                        "Access Control": "-",
                        "Host": "-",
                        "LUNs": "-",
                        "Initiator Addresses": "-",
                        "Target Address": target,
                        "Connection Type": port_type
                    })
            else:
                host_data.append({
                    "Access Control": "-",
                    "Host": "-",
                    "LUNs": "-",
                    "Initiator Addresses": "-",
                    "Target Address": "-",
                    "Connection Type": "-"
                })

        # Remove duplicates and group by all keys except "Target Address"
        def get_group_key(d):
            return tuple((k, v) for k, v in d.items() if k not in ["Target Address", "Initiator Addresses"])
        
        grouped = defaultdict(lambda: {
            "Target Address": set(),  # Using set to avoid duplicates
            "Initiator Addresses": set(),
            "original_keys": None
        })
        
        # First pass: collect unique addresses and remember key order
        for item in host_data:
            key = get_group_key(item)
            group = grouped[key]
            
            if group["original_keys"] is None:
                group["original_keys"] = list(item.keys())
            
            if "Target Address" in item:
                group["Target Address"].add(item["Target Address"])
            if "Initiator Addresses" in item:
                group["Initiator Addresses"].add(item["Initiator Addresses"])
        
        # Second pass: build merged dictionaries
        merged = []
        for key, data in grouped.items():
            new_dict = dict(key)
            original_keys = data["original_keys"]
            
            # Insert addresses in original key order
            if original_keys:
                last_key = original_keys[-1]
                last_value = new_dict.pop(last_key)
                
                if data["Target Address"]:
                    new_dict["Target Address"] = "__ ".join(sorted(data["Target Address"]))
                if data["Initiator Addresses"]:
                    new_dict["Initiator Addresses"] = "__ ".join(sorted(data["Initiator Addresses"]))
                
                new_dict[last_key] = last_value
            else:
                if data["Target Address"]:
                    new_dict["Target Address"] = "__ ".join(sorted(data["Target Address"]))
                if data["Initiator Addresses"]:
                    new_dict["Initiator Addresses"] = "__ ".join(sorted(data["Initiator Addresses"]))
            
            merged.append(new_dict)
        return merged


    except Exception as e:
        print(f"[ERROR] extract_host_info failed: {e}")
        return []

    except ValueError as e:
        print(f"Error: {e}")
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        if 'conn' in locals():
            conn.close()

def write_slot_info_sheet(writer, slot_data):
    wb = writer.book
    ws = wb.create_sheet("Slot Info")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")     # Light grey
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Light green for FC
    blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")     # Light blue for iSCSI

    slots_order = [str(i) for i in range(6, 0, -1)]  # Slots 6 to 1
    attributes = ["Port", "WWN", "Connection Type", "Speed"]
    cols_per_slot = len(attributes)

    # Prepare a dict for quick access
    slot_map = {slot['slot']: slot for slot in slot_data}

    # Write first row: slot header merged across cols_per_slot columns
    for i, slot_num in enumerate(slots_order):
        start_col = i * cols_per_slot + 1
        end_col = start_col + cols_per_slot - 1
        cell = ws.cell(row=1, column=start_col)
        cell.value = f"Slot{slot_num}"
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply light grey fill if the slot has data
        slot = slot_map.get(f"Slot{slot_num}")
        if slot and slot['total_ports'] > 0:
            for col in range(start_col, end_col + 1):
                ws.cell(row=1, column=col).fill = grey_fill
    # Write second row: port type merged across same columns, empty if missing slot
    for i, slot_num in enumerate(slots_order):
        slot_key = f"Slot{slot_num}"
        slot = slot_map.get(slot_key)
        start_col = i * cols_per_slot + 1
        end_col = start_col + cols_per_slot - 1
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = ws.cell(row=2, column=start_col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if slot is None or not slot.get('port_type'):
            cell.value = ""
        else:
            cell.value = slot['port_type'].upper()
            # Fill color by port type
            fill = green_fill if slot['port_type'].lower() == "fc" else blue_fill
            for col in range(start_col, end_col + 1):
                ws.cell(row=2, column=col).fill = fill        
    # Write third row: attribute headers
    for i, slot_num in enumerate(slots_order):
        slot_key = f"Slot{slot_num}"
        slot = slot_map.get(slot_key)
        start_col = i * cols_per_slot + 1

        if slot is None or slot['total_ports'] == 0:
            for j in range(cols_per_slot):
                ws.cell(row=3, column=start_col + j).value = ""
        else:
            for j, attr in enumerate(attributes):
                label = attr
                if attr == "Connection Type" and slot and slot.get('port_type', '').lower() == 'iscsi':
                    label = "MAC Address"
                ws.cell(row=3, column=start_col + j).value = label    # Determine max number of rows needed
    max_ports = max((slot['total_ports'] for slot in slot_map.values()), default=0)

    # Write port data (row 4 onward)
    for row_idx in range(max_ports):
        for i, slot_num in enumerate(slots_order):
            slot_key = f"Slot{slot_num}"
            slot = slot_map.get(slot_key)
            start_col = i * cols_per_slot + 1

            if slot is None or row_idx >= slot['total_ports']:
                for offset in range(cols_per_slot):
                    ws.cell(row=4 + row_idx, column=start_col + offset).value = ""
                continue

            port_info = slot['ports'].get(row_idx, {})
            port = row_idx
            wwn = port_info.get('wwn', '')

            # Use MAC address instead of connection type for iSCSI
            if slot.get('port_type', '').lower() == 'iscsi':
                connection = port_info.get('mac_address') or port_info.get('mac') or "-"
            else:
                connection = port_info.get('connection_type', '-')  # FC fallback

            speed = port_info.get('speed', '-')
            if not speed or speed.lower() == "unknown":
                speed = "-"

            # Write to sheet
            ws.cell(row=4 + row_idx, column=start_col).value = port
            ws.cell(row=4 + row_idx, column=start_col + 1).value = wwn
            ws.cell(row=4 + row_idx, column=start_col + 2).value = connection
            ws.cell(row=4 + row_idx, column=start_col + 3).value = speed

    # Adjust column widths
    for col in range(1, cols_per_slot * len(slots_order) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

if __name__ == "__main__":
    #Extract files
    about = about_info()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = "smarts.xlsx"
    dmesg_path =  os.path.join(script_dir, "dmesg.out")
    dmesg = ""
    uilog = ""
    scst_file = "scstfiledetail"
    storcli_file_path = os.path.join(script_dir, "pdlist.out")
    with open(storcli_file_path, "r", encoding="utf-8", errors="ignore") as file:
        storcli_content = file.read()

    if os.path.isfile(dmesg_path):
        dmesg = log_extract(dmesg_path)
    uilog_path =  os.path.join(script_dir, "sab-ui.out")
    if os.path.isfile(uilog_path):
        uilog = log_extract(uilog_path)

    host_data = extract_host_info()
    # Extract General Device Info
    sys_info = extract_sysinfo()
    
    smarts_file_path = os.path.join(script_dir, "smarts.txt")
        # Read the log files
    try:
        with open(smarts_file_path, "r", encoding="utf-8", errors="ignore") as file:
            smarts_content = file.read()
    except FileNotFoundError:
        print(f"Error: The required files were not found in the /SystemOverallInfo directory.")
        exit(1)
    # Extract SSD, HDD, and device info data
    ssd_data = extract_ssd_parameters(smarts_content)
    # Check if OS disks are present
    hdd_data = extract_hdd_parameters(smarts_content)
    enclosure_slot_data = []
    serial_numbers = set([disk["Serial Number"] for disk in ssd_data + hdd_data if disk["Serial Number"] != "Unknown"])
    enclosure_slot_data = extract_enclosure_slot_info(storcli_content, serial_numbers)
    slot_info = extract_slot_port_info()
    for disk in ssd_data + hdd_data:
        #Default value to avoid "N/A in excel"
        disk["En/Slot"] = ""
        for enc in enclosure_slot_data:
            if enc in disk["Serial Number"]:
                disk["Disk State"] = enclosure_slot_data[enc]["Disk State"]
                disk["Media Err"] = enclosure_slot_data[enc]["media_error_count"]
                disk["Other Err"] = enclosure_slot_data[enc]["other_error_count"]        
                disk["Shield Cnt"] = enclosure_slot_data[enc]["shield_counter"]
                disk["Predictive Failure"] = enclosure_slot_data[enc]["predictive_failure_count"]
                disk["En/Slot"] = enclosure_slot_data[enc]["enclosure_slot"]

    ssd_data = reorder_columns(ssd_data)
    hdd_data = reorder_columns(hdd_data)
    #OS Disk not found therefore we extract from pmc output
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pmc_files = glob.glob(os.path.join(script_dir, "output.txt"))

    disk_data = ssd_data + hdd_data
    device_info = extract_device_info(smarts_content, disk_data)
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Write SMART data to first sheet
        df_smart = pd.DataFrame(ssd_data + hdd_data)
    
        # Ensure empty rows are preserved in the DataFrame
        # Replace empty strings with NaN (optional, but helps with consistency)
        df_smart.replace("", pd.NA, inplace=True)

        # Write the DataFrame to the Excel file
        df_smart.to_excel(writer, sheet_name="SMART Data", index=False)
        if device_info:
            df_devinfo = pd.DataFrame(device_info)
            df_devinfo.to_excel(writer, sheet_name="Device Info", index=False)
        if host_data:
            df_host = pd.DataFrame(host_data)
            df_host.to_excel(writer, sheet_name="Host Info", index=False)
        if sys_info:
            df_host = pd.DataFrame(sys_info)
            df_host.to_excel(writer, sheet_name="General System Info", index=False)
        if slot_info:
            write_slot_info_sheet(writer, slot_info)
        if dmesg:
            df_dmesg = pd.DataFrame(dmesg)
            df_dmesg.to_excel(writer, sheet_name="dmesg log", index=False)
        if uilog:
            df_uilog = pd.DataFrame(uilog)
            df_uilog.to_excel(writer, sheet_name="sab-ui log", index=False)
        if about:
            df_about = pd.DataFrame(about)
            df_about.to_excel(writer, sheet_name="About", index=False)
    wb = load_workbook(excel_path)
    if "SMART Data" in wb.sheetnames:
        ws = wb["SMART Data"]
        for sheet_name in wb.sheetnames:
            if sheet_name != "Device Info" and sheet_name != "Slot Info":
                ws = wb[sheet_name]
                for column in range(1,7):
                    merge_cells_for_column(ws, column)
                for column in range(11,16):
                    merge_cells_for_column(ws, column)
            adjust_column_widths(ws) # Adjust column widths for all sheets
            for cell in ws[1]:
                cell.alignment = Alignment(wrap_text= True, horizontal='center')
        yellow_fill = PatternFill(start_color="ffffa6", end_color="FFFF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="de0a0a", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="bbe33d", end_color="FFFF00", fill_type="solid")
        grey_fill = PatternFill(start_color="cccccc", end_color="FFFF00", fill_type="solid")
        brown_fill = PatternFill(start_color="8d281e", end_color="FFFF00", fill_type="solid")
        deep_blue_fill = PatternFill(start_color="b8cbdf", end_color="b8cbdf", fill_type="solid")
            #Colour cells based on thresholds
        if "SMART Data" in wb.sheetnames:
            ws = wb["SMART Data"]

            # Define column indices with safe defaults
            max_cols = ws.max_column

            # Column indices (0-based for row[x] access)
            #It's better to avoid hard-coding column numbers TODO
            PARAM_COL = 6    # Column G
            THRESHOLD_COL = 7  # Column H
            VALUE_COL = 8     # Column I
            RAW_VALUE_COL = 9  # Column J
            OTHER_ERROR_COL = 12  # Column M
            SHIELD_CNT_COL = 13   # Column N
            PREDICT_FAIL_COL = 14  # Column O

            for row in ws.iter_rows(min_row=2):  # Skip header row (row 1)
                # Safely extract cells only if columns exist
                try:
                    param_value_cell = row[PARAM_COL] if PARAM_COL < len(row) else None
                    threshold_cell = row[THRESHOLD_COL] if THRESHOLD_COL < len(row) else None
                    value_cell = row[VALUE_COL] if VALUE_COL < len(row) else None
                    raw_value_cell = row[RAW_VALUE_COL] if RAW_VALUE_COL < len(row) else None
                    # Optional columns - only process if they exist
                    other_error_cell = row[OTHER_ERROR_COL] if OTHER_ERROR_COL < len(row) else None
                    shield_cnt_cell = row[SHIELD_CNT_COL] if SHIELD_CNT_COL < len(row) else None
                    predict_fail_cell = row[PREDICT_FAIL_COL] if PREDICT_FAIL_COL < len(row) else None
                except IndexError:
                    continue  # Skip rows that don't have enough columns

                param_str = param_value_cell.value if param_value_cell else None
                threshold_str = threshold_cell.value if threshold_cell else None
                value_str = value_cell.value if value_cell else None
                raw_value_str = raw_value_cell.value if raw_value_cell else None
                if other_error_cell:
                    other_error_str = other_error_cell.value
                    if other_error_str not in (None, '-', ''):
                        try:
                            other_error_value = int(other_error_str)
                            if 10 < other_error_value < 35:
                                other_error_cell.fill = yellow_fill
                            elif 36 < other_error_value < 70:
                                other_error_cell.fill = orange_fill
                            elif 70 < other_error_value < 100:
                                other_error_cell.fill = red_fill
                            elif other_error_value > 100:
                                other_error_cell.fill = brown_fill
                        except (ValueError, TypeError):
                            pass  # Silently handle conversion errors

                if shield_cnt_cell:
                    shield_cnt_str = shield_cnt_cell.value
                    if shield_cnt_str not in (None, '-', ''):
                        shield_cnt_cell.fill = red_fill

                if predict_fail_cell:
                    predict_fail_str = predict_fail_cell.value
                    if predict_fail_str not in (None, '-', ''):
                        predict_fail_cell.fill = red_fill


                if not all([param_value_cell, threshold_cell, value_cell, raw_value_cell]):
                    continue

                # We don't want to color this cell
                is_unused_rsvd = False
                if param_str == "Unused_Rsvd_Blk_Cnt_Tot":
                    is_unused_rsvd = True

                # Skip rows with missing/invalid thresholds
                if not threshold_str or threshold_str == "-":
                    continue

                try:
                    threshold_value = float(threshold_str)
                    threshold_caution = threshold_value * 1.5 if threshold_value != 0 else 30 
                    threshold_warning = threshold_value * 1.2 if threshold_value != 0 else 20
                except (ValueError, TypeError):
                    continue  # Skip non-numeric thresholds

                # Check Value or Raw Value
                compare_value = None
                if value_str not in (None, "-", ""):
                    try:
                        compare_value = float(value_str)
                        # To avoid colouring cells when value is 100
                        if compare_value == 100:
                            compare_value = 200
                    except (ValueError, TypeError):
                        pass

                if compare_value is None:  # Fallback to Raw Value
                    if raw_value_str not in (None, "-", ""):
                        try:
                            # because here raw value is smaller than threshold in normal state
                            threshold_caution = threshold_value * -0.8
                            threshold_warning = threshold_value * -0.9
                            compare_value = float(raw_value_str) * -1
                            threshold_value = threshold_value * -1
                        except (ValueError, TypeError):
                            continue  # Skip invalid values
                    else:
                        continue

                # Only color if we have the raw_value_cell and it's not unused_rsvd
                if raw_value_cell and not is_unused_rsvd:
                    # Highlight if value < threshold
                    if compare_value <= threshold_value:
                        raw_value_cell.fill = red_fill
                    elif compare_value <= threshold_warning:
                        raw_value_cell.fill = orange_fill
                    elif compare_value <= threshold_caution:
                        raw_value_cell.fill = yellow_fill

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    #Color the first row of sheets
                    for cell in ws[1]:# First row
                        cell.fill = deep_blue_fill
        wb.save(excel_path)
    print("SMART data, device info, and host info extracted and written to the Excel file with proper formatting.")
