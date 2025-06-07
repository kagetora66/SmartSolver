#By Safari, HPDS Tech Support
#===========ooOoo============
#
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import csv
import os
import glob
import sqlite3
import getpass
from zipfile import ZipFile
import subprocess
from openpyxl.styles import PatternFill
from pathlib import Path
import time
from collections import defaultdict
from collections import Counter
# Define required parameters
ssd_params = [
    "Reallocated_Sector_Ct",
    "Power_On_Hours",
    "Wear_Leveling_Count",
    "Used_Rsvd_Blk_Cnt_Tot",
    "Runtime_Bad_Block",
    "Reported_Uncorrect",
    "Hardware_ECC_Recovered",
    "Total_LBAs_Written",
    "Total Size Written (TB)"
]

hdd_params = [
    "Elements in grown defect list",
    "Total Uncorrected Errors",
    "Accumulated start-stop cycles",
    "Accumulated load-unload cycles"
]
micron_ssd_params = {
    "Raw_Read_Error_Rate": "1",
    "Reallocated_Sector_Ct": "5",
    "Reported_Uncorrect": "187",
    "Hardware_ECC_Recovered": "195",
    "Unused_Rsvd_Blk_Cnt_Tot": "180",
    "Total_LBAs_Written": "246"  # Unknown attribute at ID# 246
}

threshold_sata_ssd = {
    "Reallocated_Sector_Ct": "10",
    "Power_On_Hours": "5",
    "Wear_Leveling_Count": "10",
    "Used_Rsvd_Blk_Cnt_Tot": "10",
    "Runtime_Bad_Block": "10",
    "Reported_Uncorrect": "97",
    "Hardware_ECC_Recovered": "100",
    "Total Size Written (TB)": "6400"
} 

threshold_micron_ssd = {
    "Raw_Read_Error_Rate": "50",
    "Reallocated_Sector_Ct": "1",
    "Reported_Uncorrect": "0",
    "Hardware_ECC_Recovered": "0",
    "Unused_Rsvd_Blk_Cnt_Tot": "0",
    "Total Size Written (TB)": "6400"
    
}
smart_pattern = re.compile(
    r"(\d+)\s+([\w_]+)\s+0x[0-9a-fA-F]+\s+(\d+)\s+\d+\s+(\d+)\s+\w+-?\w*\s+\w+\s+-\s+(\d+)"
)
# Function to extract SSD parameters
def extract_ssd_parameters(log_content):
    data = []
    # Split the log into blocks for each disk.
    disk_blocks = re.findall(
        r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)",
        log_content, re.DOTALL
    )
    
    for block in disk_blocks:
        # Determine disk type based on certain strings
        is_ssd = not re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        is_sas_ssd = re.search(r"Transport protocol:\s+SAS", block, re.IGNORECASE)
        is_micron_ssd = re.search(r"Device Model:\s+Micron", block, re.IGNORECASE)
        disk_type = ""
        capacity_match = re.search(r"\[([^\]]+)\]", block)
        if capacity_match:
            tb_value = capacity_match.group(1)
            user_capacity = float(re.search(r'\d+(?:\.\d+)?', tb_value).group()) if re.search(r'\d+(?:\.\d+)?', tb_value) else 0            
        if is_ssd:
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            model_match =  re.search(r"Device Model:\s+(.+)", block, re.IGNORECASE)
            device_model = model_match.group(1) if model_match else "Unknown" 
            if is_sas_ssd:
                # --- SAS SSD extraction (existing logic) ---
                elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
                start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
                load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)
                read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
                write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
                endurance_indicator = re.search(r"SS Media used endurance indicator:\s+(\d+)", block)
                model_match_sas = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)
                device_model = model_match_sas.group(1) if model_match_sas else "Unknown" 
                brand = "Samsung"
                disk_type = "SSD SAS"
                read_error_value = int(read_error_match.group(1)) if read_error_match else 0
                write_error_value = int(write_error_match.group(1)) if write_error_match else 0
                total_uncorrected_errors = read_error_value + write_error_value
                total_uncorrrected_threshold = int(user_capacity)  * 5
                Element_in_defect_threshold = int(user_capacity)  * 100
        #We convert to str again for consistency
                threshold_sas_ssd = {
                    "Elements in grown defect list": str(Element_in_defect_threshold),
                    "Total Uncorrected Errors": str(total_uncorrrected_threshold),
                    "Accumulated start-stop cycles": "10000",
                    "Accumulated load-unload cycles": "300000"
                    }
                hdd_values = [
                    ("Elements in grown defect list", elements_grown_defect),
                    ("Total Uncorrected Errors", total_uncorrected_errors),
                    ("Accumulated start-stop cycles", start_stop_cycles),
                    ("SS Media used endurance indicator %", endurance_indicator)
                ]

                for param, match in hdd_values:
                    if isinstance(match, int):  # For Total Uncorrected Errors
                        raw_value = str(match)
                    elif match:
                        raw_value = match.group(1)
                    else:
                        continue
                    threshold = threshold_sas_ssd.get(param, "-")
                    data.append({
                        "Brand" : brand,
                        "Device Model": device_model, 
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": param,
                        "Threshold" : threshold,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
                    
            else:
                # --- Non-SAS SSD extraction ---
                # smart_pattern should be defined elsewhere to extract tuples:
                # (attr_id, attr_name, value, raw_value)
                smart_matches = smart_pattern.findall(block)
                total_lba_written = None  # To store Total_LBAs_Written when found
                disk_type = "SSD SATA"
                if is_micron_ssd:
                    brand = "Micron"
                    # Define the list of expected Micron SMART parameters.
                    micron_params = [
                        "Raw_Read_Error_Rate", 
                        "Reallocated_Sector_Ct", 
                        "Reported_Uncorrect", 
                        "Hardware_ECC_Recovered", 
                        "Unused_Rsvd_Blk_Cnt_Tot"
                    ]

                    for match in smart_matches:
                        attr_id, attr_name, value, threshold, raw_value = match
                        # Check for the known Micron parameters.
                        if attr_name in micron_params:
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": attr_name,
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                        # Detect Total_LBAs_Written by its attribute ID "246"
                        elif attr_id.strip() == "246":
                            try:
                                total_lba_written = int(raw_value)
                            except ValueError:
                                total_lba_written = None
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": "Total_LBAs_Written",
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                            
                    if total_lba_written is not None:
                        total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                        threshold_write_micron = 365 * 0.8 * 5 * user_capacity
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)",
                            "Threshold": str(round(threshold_write_micron, -3)),
                            "Value": "-",
                            "Raw Value": f"{total_size_written_tb:.2f}"
                        })
                else:
                    brand = "SAMSUNG"
                    disk_type = "SSD SATA"
                    # Existing logic for non-Micron SATA SSDs
                    for match in smart_matches:
                        attr_id, attr_name, value, threshold, raw_value = match
                        if attr_name in ssd_params: # ssd_params defined elsewhere
                            if attr_name == "Total_LBAs_Written":
                                try:
                                    total_lba_written = int(raw_value)
                                except ValueError:
                                    total_lba_written = None
                            data.append({
                                "Brand": brand,
                                "Device Model": device_model,
                                "Serial Number": serial_number,
                                "Interface": disk_type,
                                "Size": tb_value,
                                "Parameter": attr_name,
                                "Threshold": threshold,
                                "Value": value,
                                "Raw Value": raw_value
                            })
                    if total_lba_written is not None:
                        total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                        if "MZ" in device_model:
                            threshold_samsung_write = round(365 * 5 * user_capacity * 4, -3)
                        elif "850" in device_model:
                            threshold_samsung_write = 150
                        elif "860" in device_model:
                            threshold_samsung_write = 300
                        elif "870" in device_model:
                            threshold_samsung_write = 150
                        else:
                            threshold_samsung_write = 300 #undefined
                            ther
                        data.append({
                            "Brand": brand,
                            "Device Model": device_model,
                            "Serial Number": serial_number,
                            "Interface": disk_type,
                            "Size": tb_value,
                            "Parameter": "Total Size Written (TB)",
                            "Threshold": str(threshold_samsung_write),
                            "Value": "-",
                            "Raw Value": f"{total_size_written_tb:.2f}"
                        })
            
            # Add an empty row after each disk's data for readability.
            data.append({
                "Brand": "",
                "Device Model": "",
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
            for i in range(len(data) - 2):
                if (
                    data[i]["Parameter"] == "Total_LBAs_Written" and
                    data[i + 1]["Parameter"] == "Unused_Rsvd_Blk_Cnt_Tot"
                ):
                    # Swap their positions
                    data[i], data[i + 1] = data[i + 1], data[i]
            
    return data

# Function to extract HDD parameters
def extract_hdd_parameters(log_content):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    for block in disk_blocks:
        capacity_match = re.search(r"\[([^\]]+)\]", block)
        if capacity_match:
            tb_value = capacity_match.group(1)
        user_capacity = float(re.search(r'\d+(?:\.\d+)?', tb_value).group()) if re.search(r'\d+(?:\.\d+)?', tb_value) else 0
        total_uncorrrected_threshold = int(user_capacity)  * 5
        Element_in_defect_threshold = int(user_capacity)  * 100
        #We convert to str again for consistency
        threshold_hdd_sata = {
                "Elements in grown defect list": str(Element_in_defect_threshold),
                "Total Uncorrected Errors": str(total_uncorrrected_threshold),
                "Accumulated start-stop cycles": "10000",
                "Accumulated load-unload cycles": "300000"
                }
        is_sata = re.search(r'\bSATA\b', block, re.IGNORECASE)         
        is_hdd = re.search(r"Rotation Rate:\s+\d+ rpm", block, re.IGNORECASE)
        if is_hdd and not is_sata:

            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            disk_type = "HDD SAS"
            model_match =  re.search(r"Device Model:\s+(.*?)\s*$", block, re.IGNORECASE)
            model_match_hp = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)
            device_model = (
                    model_match.group(1) if model_match 
                    else model_match_hp.group(1) if model_match_hp
                    else "Unknown"
                    )
            #seagate_match =  re.search(r"Vendor:\s+(\S+)", block, re.IGNORECASE)
            if device_model.startswith("ST"):
                brand = "SEAGATE"
            elif device_model.startswith("TOSHIBA"):
                brand = "Toshiba"

  
            elements_grown_defect = re.search(r"Elements in grown defect list:\s+(\d+)", block)
            start_stop_cycles = re.search(r"Accumulated start-stop cycles:\s+(\d+)", block)
            load_unload_cycles = re.search(r"Accumulated load-unload cycles:\s+(\d+)", block)

            # Extract Total Uncorrected Errors
            read_error_match = re.search(r"read:.*?(\d+)\s*$", block, re.MULTILINE)
            write_error_match = re.search(r"write:.*?(\d+)\s*$", block, re.MULTILINE)
            
            read_error_value = int(read_error_match.group(1)) if read_error_match else 0
            write_error_value = int(write_error_match.group(1)) if write_error_match else 0
            total_uncorrected_errors = read_error_value + write_error_value

            # Sort the parameters in the desired order
            hdd_values = [
                ("Elements in grown defect list", elements_grown_defect),
                ("Total Uncorrected Errors", total_uncorrected_errors),
                ("Accumulated start-stop cycles", start_stop_cycles),
                ("Accumulated load-unload cycles", load_unload_cycles)
            ]
            
            for param, match in hdd_values:
                if isinstance(match, int):  # For Total Uncorrected Errors
                    raw_value = str(match)
                    threshold = threshold_hdd_sata.get(param, "-")
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": param,
                        "Threshold": threshold,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
                elif match:
                    raw_value = match.group(1) if hasattr(match, "group") else str(match)
                    threshold = threshold_hdd_sata.get(param, "-")
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": param,
                        "Threshold": threshold,
                        "Value": "-",
                        "Raw Value": raw_value
                    })
            # Add an empty row after each disk's data
            data.append({
                "Brand": "",
                "Device Model": "",
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
        elif is_hdd and is_sata:
            # --- SATA HDD extraction ---
            # smart_pattern should be defined elsewhere to extract tuples:
            # (attr_id, attr_name, value, raw_value)
            smart_matches = smart_pattern.findall(block) 
            total_lba_written = None  # To store Total_LBAs_Written when found    
            serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
            serial_number = serial_match.group(1) if serial_match else "Unknown"
            disk_type = "HDD SATA"
            model_match = re.search(r"Device Model:\s+(.*?)\s*$", block, re.IGNORECASE | re.MULTILINE)
            model_match_hp = re.search(r"Product:\s+(\S+)", block, re.IGNORECASE)

            #seagate_match =  re.search(r"Vendor:\s+(\S+)", block, re.IGNORECASE)
            #brand = seagate_match.group(1) if seagate_match else "HP"
            device_model = (
                    model_match.group(1) if model_match 
                    else model_match_hp.group(1) if model_match_hp
                    else "Unknown"
                    )
            if device_model.startswith("ST"):
                brand = "SEAGATE"
            elif device_model.startswith("TOSHIBA"):
                brand = "Toshiba"
            # Define the list of expected Micron SMART parameters.
            hdd_sata_params = [
                "Raw_Read_Error_Rate", 
                "Spin_Retry_Count", 
                "Reported_Uncorrect",
                "Command_Timeout", 
                "Current_Pending_Sector",
                "Offline_Uncorrectable",
                "Multi_Zone_Error_Rate"
            ]
            for match in smart_matches:
                attr_id, attr_name, value, threshold, raw_value = match
                # Check for the known SATA parameters.
                if attr_name in hdd_sata_params:
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": attr_name,
                        "Threshold": threshold,
                        "Value": value,
                        "Raw Value": raw_value
                    })
                    # Detect Total_LBAs_Written by its attribute ID "241"
                elif attr_id.strip() == "241":
                    try:
                        total_lba_written = int(raw_value)
                    except ValueError:
                        total_lba_written = None
                    data.append({
                        "Brand": brand,
                        "Device Model": device_model,
                        "Serial Number": serial_number,
                        "Interface": disk_type,
                        "Size": tb_value,
                        "Parameter": "Total_LBAs_Written",
                        "Threshold": threshold,
                        "Value": value,
                        "Raw Value": raw_value
                    })
                    
            if total_lba_written is not None:
                total_size_written_tb = total_lba_written / 2 / 1024 / 1024 / 1024
                threshold = threshold_micron_ssd.get("Total Size Written (TB)", "-")
                data.append({
                    "Brand": brand,
                    "Device Model": device_model,
                    "Serial Number": serial_number,
                    "Interface": disk_type,
                    "Size": tb_value,
                    "Parameter": "Total Size Written (TB)",
                    "Threshold": threshold,
                    "Value": "-",
                    "Raw Value": f"{total_size_written_tb:.2f}"
                    })

            data.append({
                "Brand": "",
                "Device Model": "",
                "Serial Number": "",
                "Parameter": "",
                "Value": "",
                "Raw Value": ""
            })
    return data

# Function to extract HDD device info (only if both values are found)
def extract_device_info(log_content):
    data = []
    disk_blocks = re.findall(r"=== START OF INFORMATION SECTION ===(.*?)(?==== START OF INFORMATION SECTION ===|\Z)", log_content, re.DOTALL)
    tmp_data = []
    for block in disk_blocks:
         serial_match = re.search(r"Serial Number:\s+(\S+)", block, re.IGNORECASE)
         serial_number = serial_match.group(1) if serial_match else None

         temp_match = re.search(r"Current Drive Temperature:\s+(\d+)", block)
         hours_match = re.search(r"number of hours powered up\s+=\s+([\d.]+)", block)
         #For Samsung SSDs
         temp_match_sam = re.search(r"194\s+[\w_]+\s+[\w\d]+\s+\d+\s+\d+\s+\d+\s+\w+\s+\w+\s+-\s+(\d+)", block)
         if not temp_match_sam: temp_match_sam = re.search(r"190\s+[\w_]+\s+[\w\d]+\s+\d+\s+\d+\s+\d+\s+\w+\s+\w+\s+-\s+(\d+)", block) 
         hours_match_sam = re.search(r"9\s+[\w_]+\s+[\w\d]+\s+\d+\s+\d+\s+\d+\s+\w+\s+\w+\s+-\s+(\d+)", block)
         if serial_number and temp_match:
             temperature = f"{temp_match.group(1)}"
             if hours_match:                
                 hours = hours_match.group(1)
                 tmp_data.append({
                     "Device": serial_number,
                     "Temperature": temperature,
                     "Powered Up Hours": hours
                 })
             else:
                data.append({
                     "Device": serial_number,
                     "Temperature": temperature,
                 })
         elif serial_number and temp_match_sam:
             temperature = f"{temp_match_sam.group(1)}"
             if hours_match_sam:
                 hours = hours_match_sam.group(1)
             data.append({
                 "Device": serial_number,
                 "Temperature": temperature,
                 "Powered Up Hours": hours
                 })

    data.extend(tmp_data)
    return data

# Function to extract enclosure/slot information
def extract_enclosure_slot_info(log_content, serial_numbers):
    enclosure_slot_data = {}

    # Split the log content into lines
    lines = log_content.splitlines()

    current_shield_counter = None
    current_media_error_count = None
    current_other_error_count = None
    predictive_failure_count = None
    disk_status = None
    for i, line in enumerate(lines):
        line = line.strip()

        # Capture Shield Counter
        if line.startswith("Shield Counter"):
            current_shield_counter = line.split("=")[1].strip()
            if current_shield_counter == "0":
                current_shield_counter = "-"

        # Capture Media Error Count
        elif line.startswith("Media Error Count"):
            current_media_error_count = line.split("=")[1].strip()
            if current_media_error_count == "0":
                current_media_error_count = "-"


        # Capture Other Error Count
        elif line.startswith("Other Error Count"):
            current_other_error_count = line.split("=")[1].strip()
            if current_other_error_count == "0":
                current_other_error_count = "-"

        #Capture Predictive Failure Count
        elif line.startswith("Predictive Failure Count"):
            predictive_failure_count = line.split("=")[1].strip()
            if predictive_failure_count == "0":
               predictive_failure_count = "-"

        #Capture disk state 
        
        if " UGood " in line:
            disk_status = "No Config"
        elif " Onln " in line:
            disk_status = "Operational"
        elif " DHS " in line:
            disk_status = "Hotspare"
        elif " UBad " in line:
            disk_status = "BAD UNCONFIGURED"


        # Detect Serial Number
        elif line.startswith("SN ="):
            serial = line.split("=")[1].strip()
            if serial in serial_numbers:
                # Search backward for Drive line
                for j in range(i, max(i - 20, -1), -1):
                    drive_line = lines[j].strip()
                    if drive_line.startswith("Drive /c"):
                        enclosure = drive_line.split("/e")[1].split("/")[0]
                        slot = drive_line.split("/s")[1].split()[0]

                        # Now assign all gathered info
                        enclosure_slot_data[serial] = {
                            "enclosure_slot": f"{enclosure}/{slot}",
                            "shield_counter": current_shield_counter,
                            "media_error_count": current_media_error_count,
                            "other_error_count": current_other_error_count,
                            "predictive_failure_count": predictive_failure_count,
                            "Disk State": disk_status
                        }
                        break

    return enclosure_slot_data
def extract_sysinfo():
    sysinfo_file = os.path.join(script_dir, "SystemOverallInfo", "SystemInfo.mylinux")
    version_file = os.path.join(script_dir, "version")
    pmc_file = glob.glob(os.path.join(script_dir, "output.txt"))
    sys_info = {}
    voltage_index = 1 #For separting the two power modules
    current_index = 1 #For separating the two power modules
    try:
        if pmc_file:
            # Extract versions from pmc output
            with open(pmc_file[0], "r") as file:
                content = file.read()
                basic_info = {
                    "SAB ID": re.search(r'hostname\s*:\s*(.*)$', content, re.IGNORECASE | re.MULTILINE), 
                    "SAB Version": re.search(r'#SAB version\s+([^\s]+)', content),
                    "Replication Version": re.search(r'REPLICATION VERSION:\s*VERSION=([^\s]+)', content, re.IGNORECASE),
                    "Rapidtier Version": re.search(r'Rapidtier Version:\s*([^\s]+)', content),
                    "UI Version": re.search(r'version\s*=\s*"([^"]+)"', content),
                    "CLI Version": re.search(r'CLI Version:\s*([^\s]+)', content),
                    "ROC Temp": re.search(r'ROC temperature.*\(Degree Celsius\)\s*=\s*([^\s]+)', content),
                    "CV Temp": re.search(r'Temperature\s*([^\s]+)', content),
                    "BBU Status": re.search(r'BBU Status =\s*([^\s]+)', content)
                } 

        else:
            # Extract versions from version file
            with open(version_file, "r") as file:
                content = file.read()
                basic_info = {
                    "UI Version": re.search(r'UI Version:\s*([\d.]+)', content),
                    "CLI Version": re.search(r'CLI Version:\s*([\d.]+)', content),
                    "SAB Version": re.search(r'SAB Version:\s*([\d.]+)', content)
                }

            
        for name, match in basic_info.items():
            if match:
                sys_info[name] = match.group(1)
            else:
                sys_info[name] = "Not Found"
        if "BBU Status" in sys_info:
           bbu_value = "Unknown" if sys_info["BBU Status"] is None else sys_info["BBU Status"]
           sys_info["BBU Status"] = "OK" if bbu_value == "0" else "Not OK"
        # Extract uptime and serial number from SystemInfo.mylinux
        with open(sysinfo_file, "r") as file:
            for line in file:
                uptime_match = re.search(r"up\s+(\d+)\s+days?", line)
                serial_match = re.search(r"Serial Number:\s*ZM(\S+)", line)
                voltage_match = re.search(r"Input Voltage\s*\|\s*([\d.]+)\s*V", line)
                current_match = re.search(r"Input Current \s*\|\s*([\d.]+)\s*A", line) 
                if uptime_match:
                    sys_info["Uptime (days)"] = int(uptime_match.group(1))
                    break
                else:
                    sys_info["Uptime (days)"] = 0
                if serial_match:
                     sys_info["Serial Number"] = serial_match.group(1)
                if voltage_match:
                    sys_info[f"Voltage{voltage_index}"] = float(voltage_match.group(1))
                    voltage_index += 1
                if current_match:
                    sys_info[f"Current{current_index}"] = float(current_match.group(1))
                    current_index += 1
        # Convert to list of dict format that pandas expects
        return [sys_info]
    except Exception as e:
        print(f"Error extracting system info: {e}")
        return []

# Function to extract host information from SCST configuration
def extract_host_info():
    scst_dir = "./SCST"
    db_dir = "./Database"
    # Checks for new scst file inside script directory
    new_scst_matches = glob.glob(os.path.join(script_dir, "scst.*"))
    if new_scst_matches:
        scst_files = new_scst_matches[0]
        input_file = scst_files
    else:
        scst_files = sorted(glob.glob(os.path.join(scst_dir, "scst_20*.conf")), reverse=True)
        input_file = scst_files[0] if scst_files else None
    
    if not input_file:
        print("Error: No 'scst_*.conf' files found in /SCST directory.")
        return []

    # Checks for output.txt file inside script directory 
    is_pmc = os.path.isfile(os.path.join(os.path.dirname(__file__), 'output.txt'))
    print(f"[DEBUG] SCST file used: {input_file}")
    if is_pmc:
        #print("PMC output found")
        pmc_output = "output.txt"
        target_port_type = {}
        current_wwn = None
        with open(pmc_output, "r") as file:
            lines = file.readlines()
        # We map port connection to wwn addresses 
        for line in lines:
            line = line.strip()
            if line.lower().startswith("wwn = 0x"):
                hex_str = line.split('=')[1].strip().lower().replace('0x', '')
                current_wwn = ':'.join(hex_str[i:i+2] for i in range(0, len(hex_str), 2))
            elif ('Point' in line or 'NPort' in line) and current_wwn:
                port_type = "Point to Point" if 'Point' in line else "SAN Switch"
                target_port_type[current_wwn] = port_type
                current_wwn = None
    else:
        target_port_type = {}

    try:
        with open(input_file, "r") as file:
            lines = file.readlines()

        target_blocks = []
        current_target = None
        brace_count = 0

        # Manually parse TARGET blocks
        for line in lines:
            target_match = re.match(r"\s*TARGET\s+([0-9a-fA-F:]+)\s*\{", line)
            if target_match:
                if current_target:
                    print("[WARNING] Nested TARGET found, skipping previous unfinished block.")
                current_target = {
                    "address": target_match.group(1),
                    "content": [line]
                }
                brace_count = 1
                continue

            if current_target:
                current_target["content"].append(line)
                brace_count += line.count("{")
                brace_count -= line.count("}")
                if brace_count == 0:
                    # Complete block
                    target_blocks.append((current_target["address"], "".join(current_target["content"])))
                    current_target = None

        host_data = []

        for target_address, target_body in target_blocks:
            groups = re.findall(r"GROUP\s+([\w-]+)\s*\{([\s\S]*?)\}", target_body)

            for group_name, group_content in groups:
                # Handle empty Access Control
                access_control = group_name if group_name else "-"
                
                luns = re.findall(r"LUN\s+(\d+)\s+([\w_]+)", group_content)
                initiators = re.findall(r"INITIATOR\s+([0-9a-fA-F:]+)", group_content)

                unique_luns = sorted(set(luns), key=lambda x: int(x[0]))
                unique_initiators = sorted(set(initiators))

                sab_db_file = os.path.join(db_dir, "sab.db")
                host_map = {}

                if os.path.exists(sab_db_file):
                    with sqlite3.connect(sab_db_file) as conn:
                        cursor = conn.cursor()
                        for initiator in unique_initiators:
                            cursor.execute("""
                                SELECT initiator_name, to_host_id 
                                FROM hostinitiators
                                WHERE initiator_name = ?
                            """, (initiator,))
                            to_host_row = cursor.fetchone()

                            if to_host_row:
                                to_host_id = to_host_row[1]
                                cursor.execute("""
                                    SELECT NAME 
                                    FROM host
                                    WHERE ID = ?
                                """, (to_host_id,))
                                host_row = cursor.fetchone()
                                host_map[initiator] = host_row[0] if host_row else "Not Found"
                            else:
                                host_map[initiator] = "Not Found"

                # Group initiators and LUNs by host
                host_luns_initiators = {}
                for initiator, host_name in host_map.items():
                    if host_name not in host_luns_initiators:
                        host_luns_initiators[host_name] = {
                            "luns": set(),
                            "initiators": [],
                            "target_map": {}
                        }
                    host_luns_initiators[host_name]["luns"].update(
                        [lun[1] for lun in unique_luns if lun[1] != "device_null"]
                    )
                    host_luns_initiators[host_name]["initiators"].append(initiator)
                    host_luns_initiators[host_name]["target_map"][initiator] = target_address

                for host_name, data in host_luns_initiators.items():
                    host_name = host_name if host_name else "-"
                    for initiator in sorted(data["initiators"]):
                        initiator = initiator if initiator else "-"
                        target = data["target_map"].get(initiator, "-")
                        port_type = target_port_type.get(target, "-") if target_port_type else "-"
                        
                        if not data["luns"]:
                            host_data.append({
                                "Access Control": access_control,
                                "Host": host_name,
                                "LUNs": "-",
                                "Initiator Addresses": initiator,
                                "Target Address": target,
                                "Connection Type": port_type
                            })
                        else:
                            for lun in sorted(data["luns"]):
                                host_data.append({
                                    "Access Control": access_control,
                                    "Host": host_name,
                                    "LUNs": lun,
                                    "Initiator Addresses": initiator,
                                    "Target Address": target,
                                    "Connection Type": port_type
                                })

        if not host_data:
            if target_port_type:
                for target, port_type in target_port_type.items():
                    host_data.append({
                        "Access Control": "-",
                        "Host": "-",
                        "LUNs": "-",
                        "Initiator Addresses": "-",
                        "Target Address": target,
                        "Connection Type": port_type
                    })
            else:
                host_data.append({
                    "Access Control": "-",
                    "Host": "-",
                    "LUNs": "-",
                    "Initiator Addresses": "-",
                    "Target Address": "-",
                    "Connection Type": "-"
                })

        return host_data


    except Exception as e:
        print(f"[ERROR] extract_host_info failed: {e}")
        return []

    except ValueError as e:
        print(f"Error: {e}")
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        if 'conn' in locals():
            conn.close()
#convert wwn
def convert_wwn_hex_to_colon_format(wwn_hex: str) -> str:
    """Convert WWN from hex string like '0x51402ec001c676bc' to colon-separated format."""
    hex_clean = wwn_hex.lower().replace("0x", "")
    if len(hex_clean) != 16:
        return wwn_hex  # Return as-is if not a valid WWN length
    return ":".join(hex_clean[i:i+2] for i in range(0, 16, 2))
#Extract slot number
def extract_slot_port_info():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pmc_files = glob.glob(os.path.join(script_dir, "output.txt"))

    if pmc_files:

        with open(pmc_files[0], "r") as f:
            log = f.read()

        slot_data = defaultdict(lambda: {
            'enabled_ports': [],
            'total_ports': 0,
            'port_type': '',
            'ports': {}
        })
    
        current_slot = None
        current_type = None
        current_port = None

        for line in log.splitlines():
            line = line.strip()

            # Detect slot type
            if "--------ISCSI CARDS SLOTS" in line:
                current_type = 'iscsi'
            elif "--------FC HBA CARDS SLOTS" in line:
                current_type = 'fc'

            # Match slot/port headers
            match = re.search(r'(NIC|FC) CARD in SLOT (\d+) PORT (\d+)', line)
            if match:
                current_slot = match.group(2)
                current_port = int(match.group(3))

                slot_info = slot_data[current_slot]
                slot_info['port_type'] = current_type
                slot_info['total_ports'] += 1
                slot_info['ports'][current_port] = {
                    'wwn': '',
                    'connection_type': '',
                    'speed': '-'
                }
                continue

            # FC ports
            if current_type == 'fc':
                if line.startswith("port_speed"):
                    speed = line.split("=", 1)[1].strip()
                    slot_data[current_slot]['ports'][current_port]['speed'] = speed if "Unknown" not in speed else "-"
                elif line.startswith("wwn ="):
                    wwn = line.split("=", 1)[1].strip().lower().replace("0x", "")
                    wwn_formatted = ":".join(wwn[i:i+2] for i in range(0, len(wwn), 2))
                    slot_data[current_slot]['ports'][current_port]['wwn'] = wwn_formatted
                elif line.startswith("port_type"):
                    port_type = line.split("=", 1)[1].strip()
                    if "NPort" in port_type:
                        slot_data[current_slot]['ports'][current_port]['connection_type'] = "SAN-Switch"
                    else:
                        slot_data[current_slot]['ports'][current_port]['connection_type'] = "Point-to-Point"

            # iSCSI ports
            elif current_type == 'iscsi':
                if line.startswith("speed_interface"):
                    speed = line.split("=", 1)[1].strip()
                    slot_data[current_slot]['ports'][current_port]['speed'] = speed if "Unknown" not in speed else "-"
                elif line.startswith("mac_address"):
                    mac = line.split("=", 1)[1].strip()
                    slot_data[current_slot]['ports'][current_port]['mac_address'] = mac
                elif line.startswith("iqn."):
                    iqn = line.strip()
                    slot_data[current_slot]['ports'][current_port]['wwn'] = iqn

        result = []
        for slot, info in slot_data.items():
            result.append({
                'slot': f"Slot{slot}",
                'port_type': info['port_type'],
                'total_ports': info['total_ports'],
                'ports': info['ports']
            })

        return result
#For the purpose of counting disk type and size for LOM
def consolidate_by_serial(smart_list):
    consolidated = {}

    for entry in smart_list:
        serial = entry.get("Serial Number", "").strip()
        if not serial:
            continue  # Skip empty or malformed entries

        if serial not in consolidated:
            # Initialize with static fields
            consolidated[serial] = {
                "Serial Number": serial,
                "Brand": entry.get("Brand", ""),
                "Device Model": entry.get("Device Model", ""),
                "Interface": entry.get("Interface", ""),
                "Size": entry.get("Size", "")
            }

        param = entry.get("Parameter", "").strip()
        raw_val = entry.get("Raw Value", "").strip()

        if param:
            consolidated[serial][param] = raw_val

    return list(consolidated.values())
def lom_disk(disk_dicts):
    type_size_list = []
    disk_data = consolidate_by_serial(disk_dicts)

    for d in disk_data:
        interface = d.get("Interface", "").strip()
        size = d.get("Size", "").strip()
        model = d.get("Device Model", "").strip()
        if interface and size and model:
            disk_type = interface.split()[0]  # Get "HDD" or "SSD"
            type_size_list.append((disk_type, size, model))

    counts = Counter(type_size_list)

    lom_disk_count = [
        {"Type": t, "Size": s, "Model": m, "Count": c}
        for (t, s, m), c in counts.items()
    ]
    return lom_disk_count
#Extracts full_log using a RUST program
def extractor():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    rust_binary = os.path.join(script_dir, "extractor.exe" if os.name == "nt" else "extractor")
    
    if not os.path.isfile(rust_binary):
        raise FileNotFoundError(f"Extractor binary not found at {rust_binary}")

    try:
        # Launch Rust binary (non-blocking)
        subprocess.Popen([rust_binary])
        print("Rust extractor started in background.")
        
        # Simulate work with extracted files
        print("Processing extracted files...")
        time.sleep(5)  # Replace with actual work
        while True:
            confirm = os.path.join(script_dir, "extracted_confirm")
            if os.path.isfile(confirm):
                print("Extraction Confirmed")
                response = input("Delete extracted files after finishing? (y/n): ").lower()
                if response in ['y', 'n']:
                    break
                print("Please enter 'y' or 'n'")
        
        # Create appropriate signal file
        signal_file = "delete_confirmed.flag" if response == 'y' else "delete_cancelled.flag"
        Path(signal_file).touch()

        
    except subprocess.CalledProcessError as e:
        print(f"[Python] Rust extractor failed: {e}")
        return False
# Reorder columns to make "Enclosure/Slot" the first column
def reorder_columns(data):
    return [{"En/Slot": disk.get("En/Slot", "N/A"), **disk} for disk in data]
#Extract files
script_dir = os.path.dirname(os.path.abspath(__file__))
if not os.path.isfile(os.path.join(script_dir, 'version')):
    extractor()
# Path to the smarts.mylinux file in the /SystemOverallInfo directory
smarts_file_path = os.path.join(script_dir, "SystemOverallInfo", "smarts.mylinux")
target_data = extract_host_info()

# Path to the storcli-Sall-show-all.mylinux file in the /SystemOverallInfo directory
storcli_file_path = os.path.join(script_dir, "SystemOverallInfo", "storcli-Sall-show-all.mylinux")

# Read the log files
try:
    with open(smarts_file_path, "r") as file:
        smarts_content = file.read()
    with open(storcli_file_path, "r") as file:
        storcli_content = file.read()
except FileNotFoundError:
    print(f"Error: The required files were not found in the /SystemOverallInfo directory.")
    exit(1)

# Extract SSD, HDD, and device info data
ssd_data = extract_ssd_parameters(smarts_content)
# Check if OS disks are present


hdd_data = extract_hdd_parameters(smarts_content)
device_data = extract_device_info(smarts_content)
ssd_lom = []
hdd_lom = []
if ssd_data:
    ssd_lom = lom_disk(ssd_data)
if hdd_data:
    hdd_lom = lom_disk(hdd_data)
# Extract host information
host_data = extract_host_info()
#print(host_data)
# Extract General Device Info
sys_info = extract_sysinfo()
# Extract enclosure/slot information
serial_numbers = set([disk["Serial Number"] for disk in ssd_data + hdd_data if disk["Serial Number"] != "Unknown"])
enclosure_slot_data = extract_enclosure_slot_info(storcli_content, serial_numbers)

# Add enclosure/slot information to SSD and HDD data
for disk in ssd_data + hdd_data:
    if disk["Serial Number"] in enclosure_slot_data:
        disk["Disk State"] = enclosure_slot_data[disk["Serial Number"]]["Disk State"]
        disk["Media Err"] = enclosure_slot_data[disk["Serial Number"]]["media_error_count"]
        disk["Other Err"] = enclosure_slot_data[disk["Serial Number"]]["other_error_count"]        
        disk["Shield Cnt"] = enclosure_slot_data[disk["Serial Number"]]["shield_counter"]
        disk["Predictive Failure"] = enclosure_slot_data[disk["Serial Number"]]["predictive_failure_count"]
        disk["En/Slot"] = enclosure_slot_data[disk["Serial Number"]]["enclosure_slot"]

    else:
        disk["En/Slot"] = ""
is_os = any("250 GB" in str(value) for d in ssd_data for value in d.values())
#OS Disk not found therefore we extract from pmc output
if not is_os:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pmc_files = glob.glob(os.path.join(script_dir, "output.txt"))
    if pmc_files:
        with open(pmc_files[0], "r") as f:
            log = f.read()
        ssd_os = extract_ssd_parameters(log)
        for disk in ssd_os:
            disk["En/Slot"] = "Not in RC"
        ssd_data.extend(ssd_os)

def write_slot_info_sheet(writer, slot_data):
    wb = writer.book
    ws = wb.create_sheet("Slot Info")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")     # Light grey
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Light green for FC
    blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")     # Light blue for iSCSI

    slots_order = [str(i) for i in range(6, 0, -1)]  # Slots 6 to 1
    attributes = ["Port", "WWN", "Connection Type", "Speed"]
    cols_per_slot = len(attributes)

    # Prepare a dict for quick access
    slot_map = {slot['slot']: slot for slot in slot_data}

    # Write first row: slot header merged across cols_per_slot columns
    for i, slot_num in enumerate(slots_order):
        start_col = i * cols_per_slot + 1
        end_col = start_col + cols_per_slot - 1
        cell = ws.cell(row=1, column=start_col)
        cell.value = f"Slot{slot_num}"
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply light grey fill if the slot has data
        slot = slot_map.get(f"Slot{slot_num}")
        if slot and slot['total_ports'] > 0:
            for col in range(start_col, end_col + 1):
                ws.cell(row=1, column=col).fill = grey_fill
    # Write second row: port type merged across same columns, empty if missing slot
    for i, slot_num in enumerate(slots_order):
        slot_key = f"Slot{slot_num}"
        slot = slot_map.get(slot_key)
        start_col = i * cols_per_slot + 1
        end_col = start_col + cols_per_slot - 1
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = ws.cell(row=2, column=start_col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if slot is None or not slot.get('port_type'):
            cell.value = ""
        else:
            cell.value = slot['port_type'].upper()
            # Fill color by port type
            fill = green_fill if slot['port_type'].lower() == "fc" else blue_fill
            for col in range(start_col, end_col + 1):
                ws.cell(row=2, column=col).fill = fill        
    # Write third row: attribute headers
    for i, slot_num in enumerate(slots_order):
        slot_key = f"Slot{slot_num}"
        slot = slot_map.get(slot_key)
        start_col = i * cols_per_slot + 1

        if slot is None or slot['total_ports'] == 0:
            for j in range(cols_per_slot):
                ws.cell(row=3, column=start_col + j).value = ""
        else:
            for j, attr in enumerate(attributes):
                label = attr
                if attr == "Connection Type" and slot and slot.get('port_type', '').lower() == 'iscsi':
                    label = "MAC Address"
                ws.cell(row=3, column=start_col + j).value = label    # Determine max number of rows needed
    max_ports = max((slot['total_ports'] for slot in slot_map.values()), default=0)

    # Write port data (row 4 onward)
    for row_idx in range(max_ports):
        for i, slot_num in enumerate(slots_order):
            slot_key = f"Slot{slot_num}"
            slot = slot_map.get(slot_key)
            start_col = i * cols_per_slot + 1

            if slot is None or row_idx >= slot['total_ports']:
                for offset in range(cols_per_slot):
                    ws.cell(row=4 + row_idx, column=start_col + offset).value = ""
                continue

            port_info = slot['ports'].get(row_idx, {})
            port = row_idx
            wwn = port_info.get('wwn', '')

            # Use MAC address instead of connection type for iSCSI
            if slot.get('port_type', '').lower() == 'iscsi':
                connection = port_info.get('mac_address') or port_info.get('mac') or "-"
            else:
                connection = port_info.get('connection_type', '-')  # FC fallback

            speed = port_info.get('speed', '-')
            if not speed or speed.lower() == "unknown":
                speed = "-"

            # Write to sheet
            ws.cell(row=4 + row_idx, column=start_col).value = port
            ws.cell(row=4 + row_idx, column=start_col + 1).value = wwn
            ws.cell(row=4 + row_idx, column=start_col + 2).value = connection
            ws.cell(row=4 + row_idx, column=start_col + 3).value = speed

    # Adjust column widths
    for col in range(1, cols_per_slot * len(slots_order) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


ssd_data = reorder_columns(ssd_data)
hdd_data = reorder_columns(hdd_data)

# Remove rows where "Enclosure/Slot" is "N/A"
ssd_data = [disk for disk in ssd_data]
hdd_data = [disk for disk in hdd_data]

# Create an Excel writer
excel_path = 'smart_data.xlsx'
#Create slot info
slot_info = extract_slot_port_info()

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Write SMART data to first sheet
    df_smart = pd.DataFrame(ssd_data + hdd_data)

    # Ensure empty rows are preserved in the DataFrame
    # Replace empty strings with NaN (optional, but helps with consistency)
    df_smart.replace("", pd.NA, inplace=True)

    # Write the DataFrame to the Excel file
    df_smart.to_excel(writer, sheet_name="SMART Data", index=False)

    # Write device temperature & powered-up hours to the second sheet (only if non-empty)
    if device_data:
        df_device = pd.DataFrame(device_data)
        df_device.to_excel(writer, sheet_name="Device Info", index=False)

    # Write host information to the third sheet (only if non-empty)
    if host_data:
        df_host = pd.DataFrame(host_data)
        df_host.to_excel(writer, sheet_name="Host Info", index=False)
    if sys_info:
        df_host = pd.DataFrame(sys_info)
        df_host.to_excel(writer, sheet_name="General System Info", index=False)
    if slot_info:
        write_slot_info_sheet(writer, slot_info)
    if ssd_lom or hdd_lom:
        dfs = []
        if ssd_lom:
            dfs.append(pd.DataFrame(ssd_lom))
        if hdd_lom:
            dfs.append(pd.DataFrame(hdd_lom))
        df_combined = pd.concat(dfs, ignore_index=True)
        df_combined.to_excel(writer, sheet_name="LOM", index=False)
# Open the Excel file and format it
wb = load_workbook(excel_path)
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFFF00", fill_type="solid")
#Colour cells based on thresholds
if "SMART Data" in wb.sheetnames:
    ws = wb["SMART Data"]
    for row in ws.iter_rows(min_row=2):  # Skip header row (row 1)
        # Extract relevant cells
        threshold_cell = row[5]  # Column F (Threshold)
        value_cell = row[6]      # Column G (Value)
        raw_value_cell = row[7]  # Column H (Raw Value)
        
        threshold_str = threshold_cell.value
        value_str = value_cell.value
        raw_value_str = raw_value_cell.value

        # Skip rows with missing/invalid thresholds
        if not threshold_str or threshold_str == "-":
            continue

        try:
            threshold_caution = float(threshold_str) * 1.5
            threshold_warning = float(threshold_str) * 1.2
        except (ValueError, TypeError):
            continue  # Skip non-numeric thresholds

        # Check Value or Raw Value
        compare_value = None
        if value_str not in (None, "-", ""):
            try:
                compare_value = float(value_str)
                #To avoid colouring cells when value is 100
                if compare_value == 100:
                    compare_value = 200
                    
            except (ValueError, TypeError):
                pass
        
        if compare_value is None:  # Fallback to Raw Value
            if raw_value_str not in (None, "-", ""):
                try:
                    #because here raw value is smaller than thresold in normal state
                    threshold_caution = float(threshold_str) * -0.8
                    threshold_warning = float(threshold_str) * -0.9
                    compare_value = float(raw_value_str) * -1
                   # print(threshold_warning)
                except (ValueError, TypeError):
                    continue  # Skip invalid values
            else:
                continue

        # Highlight if value < threshold
        if compare_value <= threshold_warning:
            raw_value_cell.fill = orange_fill
        elif compare_value <= threshold_caution:
            raw_value_cell.fill = yellow_fill
# Function to merge cells for a specific column
def merge_cells_for_column(ws, col_idx):
    prev_value = None
    start_row = 2  # Start from the second row (first row is headers)

    for row in range(2, ws.max_row + 1):
        current_value = ws.cell(row=row, column=col_idx).value
        if current_value == prev_value:
            continue
        if prev_value is not None and start_row is not None:
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=row-1, end_column=col_idx)
            ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")
        prev_value = current_value
        start_row = row

    if prev_value is not None and start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        ws.cell(start_row, col_idx).alignment = Alignment(vertical="center", horizontal="center")

# Function to adjust column widths automatically
def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2
deep_blue_fill = PatternFill(start_color="b8cbdf", end_color="b8cbdf", fill_type="solid")
# Format all sheets except "Device Info"
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    #Color the first row of sheets
    if sheet_name != "Slot Info":
        for cell in ws[1]:# First row
            cell.fill = deep_blue_fill
    if sheet_name not in ("Device Info", "Slot Info"):# Skip merging for "Device Info" sheet
        for column in range(1,7):
            merge_cells_for_column(ws, column)
        for column in range(11,16):
            merge_cells_for_column(ws, column)
    adjust_column_widths(ws)  # Adjust column widths for all sheets
if "Host Info" in wb.sheetnames: 
    host_info_sheet = wb["Host Info"]
    merge_cells_for_column(host_info_sheet, 4)  # Merge "Initiators" column (column 4)
    merge_cells_for_column(host_info_sheet, 5)  # Merge "Targets" column (column 5) 
    merge_cells_for_column(host_info_sheet, 6)  # Merge "Connection type" column (column 6) 
wb.save(excel_path)
print("SMART data, device info, and host info extracted and written to smart_data.xlsx with proper formatting.")
Path("python_done.flag").touch()
